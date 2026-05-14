# -*- coding: utf-8 -*-
"""
배전감리 PQ 심사 교차검증 웹 시스템
Flask 서버 - PDF OCR 추출 + Excel 교차검증 + 사전심사집계표 엑셀 다운로드
"""

import sys
import os
import re
import json
import shutil
import logging
import traceback
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
logging.basicConfig(level=logging.INFO)

from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 기존 추출기 임포트
import requests as http_requests
import warnings
warnings.filterwarnings('ignore', message='Unverified HTTPS request')

from pq_extractor import (
    analyze_company, set_current_pdf, _get_cache_dir,
    CERT_BONUS, COST_TIER
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 서버 시작 시간
SERVER_START_TIME = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# 최근 분석 결과 저장 (세션 대용)
_last_result = {}
_batch_results = []  # 일괄 분석 결과 (여러 업체)


###############################################################################
# KEEA 확인서 진위여부 조회
###############################################################################

def verify_keea_certificate(issue_no):
    """keea.or.kr 발급서류 원본확인 (발급번호로 진위여부 조회)

    Returns: True=확인됨, False=확인불가, None=조회실패
    """
    try:
        s = http_requests.Session()
        s.get('https://www.keea.or.kr/etic/won/getWON01R01.do',
              timeout=10, verify=False)
        r = s.post('https://www.keea.or.kr/etic/won/getWON01R02.do',
                   data={'srIssueNo': issue_no, 'type': ''},
                   timeout=10, verify=False)
        if r.status_code != 200:
            return None
        # fileName 값 추출: <input name="fileName" value="..."/>
        m = re.search(r'name=["\']fileName["\'][^>]*value=["\']([^"\']*)["\']', r.text)
        if not m:
            return None
        return m.group(1) != 'NO'
    except Exception as e:
        print(f"  [KEEA 조회 실패] {issue_no}: {e}")
        return None


def verify_all_keea_certs(certs):
    """여러 발급번호를 일괄 진위확인 (병렬 HTTP 요청)"""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    valid_certs = [c for c in certs if c.get("발급번호")]
    if not valid_certs:
        return []

    results = [None] * len(valid_certs)

    def _verify(idx, cert):
        issue_no = cert["발급번호"]
        verified = verify_keea_certificate(issue_no)
        status = "O" if verified else ("X" if verified is False else "?")
        print(f"  [KEEA] {cert.get('성명','')} {issue_no} → 진위여부 {status}")
        return idx, {
            "발급번호": issue_no,
            "성명": cert.get("성명", ""),
            "page": cert.get("page", -1),
            "진위확인": verified,
        }

    with ThreadPoolExecutor(max_workers=min(len(valid_certs), 4)) as executor:
        futures = [executor.submit(_verify, i, c) for i, c in enumerate(valid_certs)]
        for future in as_completed(futures):
            idx, result = future.result()
            results[idx] = result

    return results


###############################################################################
# 업체 제출 Excel 파싱
###############################################################################

def parse_company_excel(excel_path):
    """업체가 제출한 PQ심사서류 Excel 파싱"""
    result = {
        "업체명": "",
        # 책임감리원
        "책임_성명": "", "책임_등급": "",
        "책임_전기경력_년": 0, "책임_전기경력_점수": 0,
        "책임_참여분야1_년": 0, "책임_참여분야1_점수": 0,
        "책임_참여분야2_년": 0, "책임_참여분야2_점수": 0,
        "책임감리원_점수": 0,  # Sheet6 소계
        # 보조감리원
        "보조1_성명": "", "보조1_등급": "",
        "보조1_전기경력_년": 0, "보조1_전기경력_점수": 0,
        "보조1_참여분야1_점수": 0, "보조1_참여분야2_점수": 0,
        "보조감리원_점수": 0,  # Sheet6 소계
        # 비상주감리원
        "비상주_성명": "", "비상주_등급": "",
        "비상주_전기경력_년": 0, "비상주_등급_점수": 0, "비상주_전기경력_점수": 0,
        "비상주감리원_점수": 0,  # Sheet6 소계
        # 참여감리원 합계
        "참여감리원_소계": 0,
        # 유사용역
        "유사용역_적용금액": 0, "유사용역_점수": 0,
        # 기술개발 세부
        "개발실적_점수": 0, "기술투자_점수": 0, "교육실적_점수": 0,
        "기술개발_점수": 0,
        # 업무중첩 세부
        "상주중첩_점수": 0, "비상주중첩_점수": 0,
        "업무중첩_점수": 0,
        # 교체빈도 세부
        "업체교체_점수": 0, "감리원교체_점수": 0,
        "교체빈도_점수": 0,
        # 작업계획 (참고)
        "작업계획_점수": 0,
        # 가감점
        "가점_자격증": 0, "부실벌점": 0,
        "총점": 0,
        "시트목록": [],
    }

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        result["시트목록"] = wb.sheetnames
    except Exception as e:
        result["error"] = f"Excel 열기 실패: {e}"
        return result

    # --- Sheet 1: 종합평가표 (모든 점수 + 업체명) ---
    summary_sheet = None
    for name in wb.sheetnames:
        if '종합평가' in name or name.startswith('1.'):
            summary_sheet = wb[name]
            break

    if summary_sheet:
        # 업체명 검색 (상단 영역)
        for row in range(1, min(10, summary_sheet.max_row + 1)):
            for col in range(1, min(14, summary_sheet.max_column + 1)):
                val = summary_sheet.cell(row=row, column=col).value
                if val and ('㈜' in str(val) or '(주)' in str(val) or '주식회사' in str(val)):
                    result["업체명"] = str(val).strip()
                    break

        # 점수 추출: F열(column 6) = 평점
        main_section = 0
        person_section = None

        for row in range(4, min(summary_sheet.max_row + 1, 50)):
            c1 = str(summary_sheet.cell(row=row, column=1).value or '').strip()
            c3 = str(summary_sheet.cell(row=row, column=3).value or '').strip()
            c4 = str(summary_sheet.cell(row=row, column=4).value or '').strip()
            c6 = _safe_num(summary_sheet.cell(row=row, column=6).value)

            # ── 대분류 감지 ──
            if '1.' in c1 and '감리원' in c1:
                main_section = 1
                person_section = None
                result['참여감리원_소계'] = c6
                continue
            elif '2.' in c1 and '용역' in c1:
                main_section = 2
                person_section = None
                result['유사용역_점수'] = c6
                continue
            elif '3.' in c1 and '신용' in c1:
                main_section = 3
                person_section = None
                continue
            elif ('4.' in c1 and ('기술' in c1 or '투자' in c1)):
                main_section = 4
                person_section = None
                result['기술개발_점수'] = c6
                continue
            elif '5.' in c1 and '중첩' in c1:
                main_section = 5
                person_section = None
                result['업무중첩_점수'] = c6
                continue
            elif '6.' in c1 and '교체' in c1:
                main_section = 6
                person_section = None
                result['교체빈도_점수'] = c6
                continue
            elif '7.' in c1 and '작업' in c1:
                main_section = 7
                person_section = None
                result['작업계획_점수'] = c6
                continue
            elif '8.' in c1 or ('가산' in c1 and '감점' in c1):
                main_section = 8
                person_section = None
                continue
            elif '합' in c1 and '계' in c1:
                result['총점'] = c6
                continue

            # ── 1. 참여감리원 세부 ──
            if main_section == 1:
                if '책임감리원' in c3:
                    person_section = '책임'
                elif '보조감리원' in c3:
                    person_section = '보조'
                elif '비상주' in c3:
                    person_section = '비상주'

                if person_section == '책임':
                    if '(2)' in c4: result['책임_전기경력_점수'] = c6
                    elif '(3-1)' in c4: result['책임_참여분야1_점수'] = c6
                    elif '(3-2)' in c4: result['책임_참여분야2_점수'] = c6
                elif person_section == '보조':
                    if '(2)' in c4: result['보조1_전기경력_점수'] = c6
                    elif '(3-1)' in c4: result['보조1_참여분야1_점수'] = c6
                    elif '(3-2)' in c4: result['보조1_참여분야2_점수'] = c6
                elif person_section == '비상주':
                    if '(1)' in c4: result['비상주_등급_점수'] = c6
                    elif '(2)' in c4: result['비상주_전기경력_점수'] = c6

            # ── 4. 기술개발 세부 ──
            elif main_section == 4:
                if '개발실적' in c3: result['개발실적_점수'] = c6
                elif '투자실적' in c3: result['기술투자_점수'] = c6
                elif '교육실적' in c3: result['교육실적_점수'] = c6

            # ── 5. 업무중첩 세부 ──
            elif main_section == 5:
                if '상주' in c3 and '비상주' not in c3: result['상주중첩_점수'] = c6
                elif '비상주' in c3: result['비상주중첩_점수'] = c6

            # ── 6. 교체빈도 세부 ──
            elif main_section == 6:
                if '업체' in c3: result['업체교체_점수'] = c6
                elif '감리원' in c3: result['감리원교체_점수'] = c6

            # ── 8. 가감점 세부 ──
            elif main_section == 8:
                if '자격' in c3: result['가점_자격증'] = c6
                elif '부실' in c3 or '벌점' in c3: result['부실벌점'] = c6

        # 개인별 소계 계산
        result['책임감리원_점수'] = result['책임_전기경력_점수'] + result['책임_참여분야1_점수'] + result['책임_참여분야2_점수']
        result['보조감리원_점수'] = result['보조1_전기경력_점수'] + result['보조1_참여분야1_점수'] + result['보조1_참여분야2_점수']
        result['비상주감리원_점수'] = result['비상주_등급_점수'] + result['비상주_전기경력_점수']

    # --- Sheet 2: 세부항목평가표 ---
    detail_sheet = None
    for name in wb.sheetnames:
        if '세부항목' in name or '2.' in name:
            detail_sheet = wb[name]
            break

    if detail_sheet:
        for row in range(1, min(50, detail_sheet.max_row + 1)):
            for col in range(1, min(15, detail_sheet.max_column + 1)):
                val = detail_sheet.cell(row=row, column=col).value
                if val and '책임감리원' in str(val):
                    _find_personnel_in_sheet(detail_sheet, row, result, "책임")
                    break
            if result["책임_성명"]:
                break

        for row in range(1, min(100, detail_sheet.max_row + 1)):
            for col in range(1, min(15, detail_sheet.max_column + 1)):
                val = detail_sheet.cell(row=row, column=col).value
                if val and '보조감리원' in str(val) and '1' in str(val):
                    _find_personnel_in_sheet(detail_sheet, row, result, "보조1")
                    break
            if result["보조1_성명"]:
                break

        for row in range(1, min(150, detail_sheet.max_row + 1)):
            for col in range(1, min(15, detail_sheet.max_column + 1)):
                val = detail_sheet.cell(row=row, column=col).value
                if val and '비상주' in str(val):
                    _find_personnel_in_sheet(detail_sheet, row, result, "비상주")
                    break
            if result["비상주_성명"]:
                break

        _extract_career_from_detail(detail_sheet, result)

    # --- Sheet 4/5: 유사용역수행실적 ---
    for name in wb.sheetnames:
        if '유사용역' in name and '준공' in name:
            ws = wb[name]
            for row in range(1, min(30, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val and '누계' in str(val):
                        for c2 in range(col + 1, min(col + 5, ws.max_column + 1)):
                            v2 = ws.cell(row=row, column=c2).value
                            if isinstance(v2, (int, float)) and v2 > 10000:
                                result["유사용역_적용금액"] = int(v2)
                                break
            break

    wb.close()
    return result


def _safe_num(val):
    """셀 값을 숫자로 안전하게 변환"""
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def _find_personnel_in_sheet(ws, start_row, result, prefix):
    """시트에서 감리원 성명/등급 찾기"""
    grades = ['특급', '고급', '중급', '초급']
    for row in range(start_row, min(start_row + 5, ws.max_row + 1)):
        for col in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if not val:
                continue
            val_str = str(val).strip()
            if val_str in grades:
                result[f"{prefix}_등급"] = val_str
            if re.match(r'^[가-힣]{2,4}$', val_str) and not result.get(f"{prefix}_성명"):
                if val_str not in ['소계', '합계', '배점', '평점', '평가', '등급',
                                   '특급', '고급', '중급', '초급', '책임', '보조', '비상주',
                                   '감리원', '감리', '전기', '토목', '경력', '분야']:
                    result[f"{prefix}_성명"] = val_str


def _extract_career_from_detail(ws, result):
    """세부항목평가표에서 감리원별 경력 년수와 평점 추출"""
    section = None  # '책임', '보조', '비상주'
    for row in range(1, min(ws.max_row + 1, 30)):
        c1 = str(ws.cell(row=row, column=1).value or '').strip()
        c4 = str(ws.cell(row=row, column=4).value or '').strip()
        c6 = str(ws.cell(row=row, column=6).value or '').strip()
        c9 = ws.cell(row=row, column=9).value  # 평점 열

        # 섹션 판별
        if '책임감리원' in c1:
            section = '책임'
        elif '보조감리원' in c1:
            section = '보조'
        elif '비상주' in c1 and ('감리원' in c1 or '감리' in c1):
            section = '비상주'
        elif '유사용역' in c1 or '합 계' in c1:
            section = None

        if not section:
            continue

        m_years = re.search(r'(\d+\.?\d*)\s*년', c6)

        if section == '책임':
            if '전기분야' in c4 and '경력' in c4:
                if m_years and result["책임_전기경력_년"] == 0:
                    result["책임_전기경력_년"] = float(m_years.group(1))
                if _safe_num(c9) > 0:
                    result["책임_전기경력_점수"] = _safe_num(c9)
            elif '3-1' in c4 or ('참여분야' in c4 and '3-2' not in c4 and '소' not in c4):
                if m_years and result["책임_참여분야1_년"] == 0:
                    result["책임_참여분야1_년"] = float(m_years.group(1))
                if _safe_num(c9) > 0:
                    result["책임_참여분야1_점수"] = _safe_num(c9)
            elif '3-2' in c4:
                if m_years and result["책임_참여분야2_년"] == 0:
                    result["책임_참여분야2_년"] = float(m_years.group(1))
                if _safe_num(c9) > 0:
                    result["책임_참여분야2_점수"] = _safe_num(c9)
        elif section == '보조':
            if '전기분야' in c4 and '경력' in c4:
                if m_years and result["보조1_전기경력_년"] == 0:
                    result["보조1_전기경력_년"] = float(m_years.group(1))
                if _safe_num(c9) > 0:
                    result["보조1_전기경력_점수"] = _safe_num(c9)
            elif '3-1' in c4:
                if _safe_num(c9) > 0:
                    result["보조1_참여분야1_점수"] = _safe_num(c9)
            elif '3-2' in c4:
                if _safe_num(c9) > 0:
                    result["보조1_참여분야2_점수"] = _safe_num(c9)
        elif section == '비상주':
            if '등급' in c4 or '등 급' in c4:
                if _safe_num(c9) > 0:
                    result["비상주_등급_점수"] = _safe_num(c9)
            elif '전기분야' in c4 and '경력' in c4:
                if m_years and result["비상주_전기경력_년"] == 0:
                    result["비상주_전기경력_년"] = float(m_years.group(1))
                if _safe_num(c9) > 0:
                    result["비상주_전기경력_점수"] = _safe_num(c9)


###############################################################################
# 교차검증
###############################################################################

def _add_criteria(item, criteria_value, criteria_basis="", pdf_val=None, tolerance=0.1):
    """비교 항목 dict에 심사기준 계산값 필드를 추가한다."""
    item["criteria_value"] = criteria_value
    item["criteria_basis"] = criteria_basis
    if criteria_value is not None and pdf_val is not None:
        pv = _safe_num(pdf_val)
        cv = _safe_num(criteria_value)
        item["criteria_match"] = abs(pv - cv) <= tolerance
    else:
        item["criteria_match"] = None
    return item


def cross_verify(pdf_data, excel_data):
    """PDF 추출 결과와 Excel 제출 데이터 교차검증 (심사기준 계산값 포함)"""
    items = []
    cs = pdf_data.get("criteria_scores", {})  # 기준표 계산 결과

    # 책임감리원 등급 (텍스트 — 기준 계산값 없음)
    item = _compare_item("참여감리원", "책임감리원 등급",
        pdf_data.get("책임_등급", ""), excel_data.get("책임_등급", ""), severity="error")
    _add_criteria(item, None)
    items.append(item)

    # 책임감리원 성명 (텍스트)
    item = _compare_item("참여감리원", "책임감리원 성명",
        pdf_data.get("책임_성명", ""), excel_data.get("책임_성명", ""), severity="warning")
    _add_criteria(item, None)
    items.append(item)

    # 책임 전기분야 경력(년)
    item = _compare_numeric("참여감리원", "책임 전기분야경력(년)",
        pdf_data.get("책임_전기경력_년", 0), excel_data.get("책임_전기경력_년", 0),
        tolerance=0.5, severity="warning")
    _add_criteria(item, None)
    items.append(item)

    # 책임 전기분야 점수 — 기준계산값 비교
    item = _compare_numeric("참여감리원", "책임 전기분야 점수",
        pdf_data.get("책임_전기경력_점수", 0), excel_data.get("책임_전기경력_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("책임_전기경력_계산점수"),
        cs.get("책임_전기경력_근거", ""),
        pdf_data.get("책임_전기경력_점수", 0))
    items.append(item)

    # 책임 참여분야1 점수 — 기준계산값 비교
    item = _compare_numeric("참여감리원", "책임 참여분야1 점수",
        pdf_data.get("책임_배전경력1_점수", 0), excel_data.get("책임_참여분야1_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("책임_배전경력1_계산점수"),
        cs.get("책임_배전경력1_근거", ""),
        pdf_data.get("책임_배전경력1_점수", 0))
    items.append(item)

    # 책임 참여분야2 점수 — 기준계산값 비교
    item = _compare_numeric("참여감리원", "책임 참여분야2 점수",
        pdf_data.get("책임_배전경력2_점수", 0), excel_data.get("책임_참여분야2_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("책임_배전경력2_계산점수"),
        cs.get("책임_배전경력2_근거", ""),
        pdf_data.get("책임_배전경력2_점수", 0))
    items.append(item)

    # 보조감리원 등급 (텍스트)
    item = _compare_item("참여감리원", "보조감리원 등급",
        pdf_data.get("보조_등급", ""), excel_data.get("보조1_등급", ""), severity="warning")
    _add_criteria(item, None)
    items.append(item)

    # 보조감리원 전기분야 점수
    item = _compare_numeric("참여감리원", "보조 전기분야 점수",
        pdf_data.get("보조_전기경력_점수", 0), excel_data.get("보조1_전기경력_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("보조_전기경력_계산점수"),
        cs.get("보조_전기경력_근거", ""),
        pdf_data.get("보조_전기경력_점수", 0))
    items.append(item)

    # 보조감리원 참여분야1 점수
    item = _compare_numeric("참여감리원", "보조 참여분야1 점수",
        pdf_data.get("보조_참여분야1_점수", 0), excel_data.get("보조1_참여분야1_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("보조_참여1_계산점수"),
        cs.get("보조_참여1_근거", ""),
        pdf_data.get("보조_참여분야1_점수", 0))
    items.append(item)

    # 보조감리원 참여분야2 점수
    item = _compare_numeric("참여감리원", "보조 참여분야2 점수",
        pdf_data.get("보조_참여분야2_점수", 0), excel_data.get("보조1_참여분야2_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("보조_참여2_계산점수"),
        cs.get("보조_참여2_근거", ""),
        pdf_data.get("보조_참여분야2_점수", 0))
    items.append(item)

    # 비상주감리원 등급 — 기준계산값(등급→점수) 추가
    item = _compare_item("참여감리원", "비상주감리원 등급",
        pdf_data.get("비상주_등급", ""), excel_data.get("비상주_등급", ""), severity="error")
    nonres_cv = cs.get("비상주_등급_계산점수")
    _add_criteria(item,
        f"{cs.get('비상주_등급_근거','')} → {nonres_cv}점" if nonres_cv is not None else None,
        cs.get("비상주_등급_근거", ""))
    items.append(item)

    # 비상주감리원 전기분야 점수
    item = _compare_numeric("참여감리원", "비상주 전기분야 점수",
        pdf_data.get("비상주_전기경력_점수", 0), excel_data.get("비상주_전기경력_점수", 0),
        tolerance=0.1, severity="error")
    _add_criteria(item,
        cs.get("비상주_전기경력_계산점수"),
        cs.get("비상주_전기경력_근거", ""),
        pdf_data.get("비상주_전기경력_점수", 0))
    items.append(item)

    # 유사용역 점수 — 기준계산값 비교
    item = _compare_numeric("유사용역", "유사용역 수행실적 점수",
        pdf_data.get("유사용역_점수", 0),
        excel_data.get("유사용역_점수", 0), tolerance=0.5, severity="error")
    _add_criteria(item,
        cs.get("유사용역_계산점수"),
        cs.get("유사용역_근거", ""),
        pdf_data.get("유사용역_점수", 0),
        tolerance=0.5)
    items.append(item)

    # 기술개발 — 개발실적 (기간비율 × 타입 독립 계산)
    item = _compare_numeric("기술개발", "가. 개발실적",
        pdf_data.get("개발실적_점수", pdf_data.get("기술개발_점수", 0)),
        excel_data.get("개발실적_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("개발실적_계산점수"),
        cs.get("개발실적_근거", ""),
        pdf_data.get("개발실적_점수", 0),
        tolerance=0.5)
    items.append(item)

    # 기술개발 — 기술투자실적 (A÷B% 독립 계산)
    item = _compare_numeric("기술개발", "나. 기술투자실적",
        pdf_data.get("기술투자_점수", 0),
        excel_data.get("기술투자_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("기술투자_계산점수"),
        cs.get("기술투자_근거", ""),
        pdf_data.get("기술투자_점수", 0),
        tolerance=0.5)
    items.append(item)

    # 기술개발 — 교육실적
    item = _compare_numeric("기술개발", "다. 교육실적",
        pdf_data.get("교육실적_기재점수", 0),
        excel_data.get("교육실적_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("교육실적_계산점수"),
        cs.get("교육실적_근거", ""),
        pdf_data.get("교육실적_기재점수", 0),
        tolerance=0.1)
    items.append(item)

    # 업무중첩 — 가. 상주감리원
    item = _compare_numeric("업무중첩", "가. 상주감리원",
        pdf_data.get("상주중첩_기재점수", 0),
        excel_data.get("상주중첩_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("상주중첩_계산점수"),
        cs.get("상주중첩_근거", ""),
        pdf_data.get("상주중첩_기재점수", 0), tolerance=0.1)
    items.append(item)

    # 업무중첩 — 나. 비상주감리원
    item = _compare_numeric("업무중첩", "나. 비상주감리원",
        pdf_data.get("비상주중첩_기재점수", 0),
        excel_data.get("비상주중첩_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("비상주중첩_계산점수"),
        cs.get("비상주중첩_근거", ""),
        pdf_data.get("비상주중첩_기재점수", 0), tolerance=0.1)
    items.append(item)

    # 교체빈도 — 가. 감리업체
    item = _compare_numeric("교체빈도", "가. 감리업체",
        pdf_data.get("업체교체_기재점수", 0),
        excel_data.get("업체교체_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("업체교체_계산점수"),
        cs.get("업체교체_근거", ""),
        pdf_data.get("업체교체_기재점수", 0), tolerance=0.1)
    items.append(item)

    # 교체빈도 — 나. 참여감리원
    item = _compare_numeric("교체빈도", "나. 참여감리원",
        pdf_data.get("감리원교체_기재점수", 0),
        excel_data.get("감리원교체_점수", 0), tolerance=0.5, severity="warning")
    _add_criteria(item,
        cs.get("감리원교체_계산점수"),
        cs.get("감리원교체_근거", ""),
        pdf_data.get("감리원교체_기재점수", 0), tolerance=0.1)
    items.append(item)

    # 자격증 가점 — 기준계산값 비교
    item = _compare_numeric("가점감점", "자격증 가점",
        pdf_data.get("가점_자격증", 0), excel_data.get("가점_자격증", 0),
        tolerance=0.1, severity="warning")
    _add_criteria(item,
        cs.get("가점_계산점수"),
        cs.get("가점_근거", ""),
        pdf_data.get("가점_자격증", 0))
    items.append(item)

    # 부실벌점
    item = _compare_numeric("가점감점", "부실벌점",
        0, excel_data.get("부실벌점", 0),
        tolerance=0.1, severity="warning")
    _add_criteria(item,
        cs.get("부실벌점_계산점수"),
        cs.get("부실벌점_근거", ""),
        0, tolerance=0.1)
    items.append(item)

    # 총점 (신용도 제외)
    item = _compare_numeric("합계", "총점 (신용도 제외)",
        _calc_pdf_total(pdf_data), excel_data.get("총점", 0),
        tolerance=1.0, severity="error")
    _add_criteria(item, None)
    items.append(item)

    matched = sum(1 for i in items if i["match"])
    errors = sum(1 for i in items if i["severity"] == "error" and not i["match"])
    warnings = sum(1 for i in items if i["severity"] == "warning" and not i["match"])

    disq_reasons = []
    chief_grade = pdf_data.get("책임_등급") or excel_data.get("책임_등급", "")
    if chief_grade and chief_grade not in ["특급", "고급", "중급"]:
        disq_reasons.append(f"책임감리원 등급 미달({chief_grade})")
    nonres_grade = pdf_data.get("비상주_등급") or excel_data.get("비상주_등급", "")
    if nonres_grade and nonres_grade not in ["특급", "고급"]:
        disq_reasons.append(f"비상주감리원 등급 미달({nonres_grade})")

    return {
        "items": items,
        "summary": {
            "total_items": len(items),
            "matched": matched,
            "errors": errors,
            "warnings": warnings,
            "pass_fail": "실격" if disq_reasons else "적격",
            "disq_reasons": disq_reasons,
            "pdf_total": _calc_pdf_total(pdf_data),
            "excel_total": excel_data.get("총점", 0),
        }
    }


def _compare_item(category, item_name, pdf_val, excel_val, severity="warning"):
    pdf_str = str(pdf_val).strip() if pdf_val else ""
    excel_str = str(excel_val).strip() if excel_val else ""
    match = (pdf_str == excel_str) or (not pdf_str and not excel_str)
    if not pdf_str and excel_str:
        note = "PDF 추출 실패"
        match = False
        severity = "warning"
    elif pdf_str and not excel_str:
        note = "Excel 미기입"
        match = False
    elif not match:
        note = "불일치"
    else:
        note = ""
    return {
        "category": category, "item": item_name,
        "pdf_value": pdf_str or "(미확인)", "excel_value": excel_str or "(미기입)",
        "match": match, "note": note, "severity": "ok" if match else severity,
    }


def _compare_numeric(category, item_name, pdf_val, excel_val, tolerance=0, severity="warning"):
    pv = _safe_num(pdf_val)
    ev = _safe_num(excel_val)
    if pv == 0 and ev == 0:
        match, note = True, ""
    elif pv == 0 and ev > 0:
        match, note, severity = False, "PDF 추출 실패", "warning"
    elif abs(pv - ev) <= tolerance:
        match, note = True, ""
    else:
        match = False
        note = f"차이: {round(pv - ev, 2):+.2f}"
    return {
        "category": category, "item": item_name,
        "pdf_value": pv if pv != 0 else "(미확인)",
        "excel_value": ev if ev != 0 else "(미기입)",
        "match": match, "note": note, "severity": "ok" if match else severity,
    }


def _calc_pdf_total(pdf_data):
    """PDF 추출 데이터 기반 총점 계산 (신용도/작업계획 제외)"""
    total = 0
    total += _safe_num(pdf_data.get("참여감리원_소계", 0))
    total += _safe_num(pdf_data.get("유사용역_점수", 0))
    total += _safe_num(pdf_data.get("기술개발_점수", 0))
    total += _safe_num(pdf_data.get("업무중첩_점수", 0))
    total += _safe_num(pdf_data.get("교체빈도_점수", 0))
    total += _safe_num(pdf_data.get("가점_자격증", 0))
    return round(total, 2)


###############################################################################
# 사전심사집계표 엑셀 생성
###############################################################################

def export_evaluation_sheet(results_list, output_path, template_path=None):
    if template_path and os.path.exists(template_path):
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "사전심사집계표"
        _create_template_header(ws)

    start_row = 6
    for idx, data in enumerate(results_list):
        row = start_row + idx
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=2, value=data.get("사업자번호", ""))
        ws.cell(row=row, column=3, value=data.get("업체명", ""))
        ws.cell(row=row, column=8, value=data.get("참여감리원_점수", 0))
        ws.cell(row=row, column=9, value=data.get("유사용역_점수", 0))
        ws.cell(row=row, column=10, value=data.get("기술개발_점수", 0))
        ws.cell(row=row, column=11, value=data.get("업무중첩_점수", 0))
        ws.cell(row=row, column=12, value=data.get("교체빈도_점수", 0))
        ws.cell(row=row, column=13, value=data.get("작업계획_점수", 0))
        ws.cell(row=row, column=14, value=data.get("가감점", 0))
        ws.cell(row=row, column=15, value=f"=SUM(H{row}:N{row})")
        if data.get("특이사항"):
            ws.cell(row=row, column=17, value=data["특이사항"])

    wb.save(output_path)
    return output_path


def _create_template_header(ws):
    ws.merge_cells('A1:O1')
    ws['A1'] = '사업수행능력 세부평가기준(PQ) 내역'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = '공고명 :'
    headers_score = {'H': 50, 'I': 10, 'J': 10, 'K': 10, 'L': 5, 'M': 5, 'N': '+1~-5', 'O': 90}
    for col_letter, val in headers_score.items():
        ws[f'{col_letter}2'] = val
    ws['A3'] = '공고일 : 2026.03.09'
    headers_4 = {'A': '순번', 'B': '사업자번호', 'C': '업체명', 'D': '분담비율',
                  'E': '사업자번호', 'F': '업체명', 'G': '분담비율'}
    for col, val in headers_4.items():
        ws[f'{col}4'] = val
        ws[f'{col}4'].font = Font(name='맑은 고딕', bold=True, size=9)
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('H4:O4')
    ws['H4'] = '사업수행능력 점수(공사부서 평가 : 신용도 제외)'
    ws['H4'].font = Font(name='맑은 고딕', bold=True, size=9)
    ws['H4'].alignment = Alignment(horizontal='center')
    headers_5 = {'H': '참여\n감리원', 'I': '유사용역\n수행실적', 'J': '기술개발\n및 투자실적',
                  'K': '업무\n중첩도', 'L': '교체\n빈도', 'M': '작업계획\n및 기법',
                  'N': '가점·감점', 'O': '소계'}
    for col, val in headers_5.items():
        ws[f'{col}5'] = val
        ws[f'{col}5'].font = Font(name='맑은 고딕', size=8)
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    widths = {'A': 5, 'B': 14, 'C': 18, 'D': 7, 'E': 14, 'F': 18, 'G': 7,
              'H': 8, 'I': 10, 'J': 10, 'K': 8, 'L': 7, 'M': 10, 'N': 8, 'O': 7}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


###############################################################################
# Flask 에러 핸들러 (항상 JSON 반환)
###############################################################################

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "파일 크기가 너무 큽니다 (최대 200MB)"}), 413

@app.errorhandler(500)
def internal_error(e):
    return jsonify({"error": "서버 내부 오류", "detail": str(e)}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    app.logger.error(f"Unhandled: {e}\n{traceback.format_exc()}")
    return jsonify({"error": str(e)}), 500


###############################################################################
# Flask 라우트
###############################################################################

@app.route('/')
def index():
    return render_template('index.html', server_time=SERVER_START_TIME)


@app.route('/api/analyze', methods=['POST'])
def api_analyze():
    """PDF + Excel 업로드 → 분석 → 교차검증 결과 JSON"""
    global _last_result, _batch_results

    # 첫 업체 분석 시 배치 초기화
    is_batch_start = request.form.get('batch_index', '')
    if is_batch_start == '0':
        _batch_results = []

    pdf_file = request.files.get('pdf_file')
    excel_file = request.files.get('excel_file')
    bidding_date = request.form.get('bidding_date', '2026-03-09')
    cost_tier = request.form.get('cost_tier', '10억미만')

    if not pdf_file:
        return jsonify({"error": "PDF 파일을 업로드해주세요."}), 400

    pdf_basename = os.path.basename(pdf_file.filename)
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_basename)
    pdf_file.save(pdf_path)

    excel_path = None
    if excel_file and excel_file.filename:
        excel_basename = os.path.basename(excel_file.filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_basename)
        excel_file.save(excel_path)

    try:
        # 1. PDF 분석
        pdf_data = analyze_company(pdf_path, bidding_date, cost_tier)

        # 2. Excel 파싱
        excel_data = {}
        if excel_path:
            excel_data = parse_company_excel(excel_path)

        # 3. 교차검증
        verification = cross_verify(pdf_data, excel_data)

        # 3.5 KEEA 확인서 진위여부 조회
        keea_certs = pdf_data.get("keea_certs", [])
        keea_results = []
        if keea_certs:
            print("[KEEA] 확인서 진위여부 조회 중...")
            keea_results = verify_all_keea_certs(keea_certs)

        # 4. 평가표용 데이터
        eval_data = {
            "업체명": pdf_data.get("업체명") or excel_data.get("업체명", ""),
            "사업자번호": "",
            "참여감리원_점수": pdf_data.get("참여감리원_소계", 0) or excel_data.get("참여감리원_소계", 0),
            "유사용역_점수": pdf_data.get("유사용역_점수", 0) or excel_data.get("유사용역_점수", 0),
            "기술개발_점수": pdf_data.get("기술개발_점수", 0) or excel_data.get("기술개발_점수", 0),
            "업무중첩_점수": pdf_data.get("업무중첩_점수", 0) or excel_data.get("업무중첩_점수", 0),
            "교체빈도_점수": pdf_data.get("교체빈도_점수", 0) or excel_data.get("교체빈도_점수", 0),
            "가감점": pdf_data.get("가점_자격증", 0) or (excel_data.get("가점_자격증", 0) - excel_data.get("부실벌점", 0)),
            "총점": pdf_data.get("업체제출_총점", 0) or excel_data.get("총점", 0),
        }

        _last_result = {
            "pdf_data": pdf_data,
            "excel_data": excel_data,
            "verification": verification,
            "eval_data": eval_data,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        # 배치 결과에 추가
        _batch_results.append(eval_data)

        return jsonify({
            "success": True,
            "company_name": eval_data["업체명"],
            "pdf_data": pdf_data,
            "excel_data": excel_data,
            "verification": verification,
            "eval_data": eval_data,
            "keea_results": keea_results,
        })

    except Exception as e:
        app.logger.error(f"분석 오류: {e}\n{traceback.format_exc()}")
        return jsonify({
            "error": str(e),
            "traceback": traceback.format_exc()
        }), 500


@app.route('/api/download-excel')
def api_download_excel():
    """사전심사집계표 양식 엑셀 다운로드 (전체 업체)"""
    global _batch_results

    if not _batch_results:
        return jsonify({"error": "먼저 분석을 실행해주세요."}), 400

    template_path = os.path.join(os.path.dirname(__file__), "pq심사내역 평가표_(양식).xlsx")
    output_path = os.path.join(app.config['UPLOAD_FOLDER'],
                               f"PQ심사결과_{len(_batch_results)}개업체.xlsx")

    try:
        export_evaluation_sheet(
            _batch_results,
            output_path,
            template_path=template_path if os.path.exists(template_path) else None
        )
        return send_file(output_path, as_attachment=True,
                         download_name=os.path.basename(output_path))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


###############################################################################
# 메인
###############################################################################

if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("  배전감리 PQ 심사 교차검증 시스템")
    print("  http://localhost:5002")
    print("=" * 60 + "\n")
    app.run(host='0.0.0.0', port=5002, debug=False)
