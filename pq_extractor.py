# -*- coding: utf-8 -*-
"""
배전감리 PQ 심사 데이터 추출 시스템
한국전력공사 배전건설부 - 감리용역 적격심사 PQ 자동 추출

사용법: python pq_extractor.py <업체PDF경로> [입찰공고일] [총예정전기공사비구간]
예시: python pq_extractor.py "증빙/(주)성문기술단.pdf" 2026-03-09 10억미만
"""

import sys
import os
import re
import json
from datetime import datetime, timedelta
from pathlib import Path

# Windows 한글 출력 설정
sys.stdout.reconfigure(encoding='utf-8')

# PDF/OCR
import fitz  # PyMuPDF
import easyocr

# Excel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

###############################################################################
# 설정
###############################################################################
BIDDING_DATE = "2026-03-09"  # 입찰공고일 (기본값)
COST_TIER = "10억미만"  # 총 예정전기공사비 구간

###############################################################################
# 경력 점수 기준표 (공사비 구간 + 등급 기반)
###############################################################################
# 공사비 구간별 기본 경력년수 임계값 (만점 기준)
# 100억이상→7, 50~100억→6, 10~50억→5, 10억미만→4
COST_TIER_BASE = {"100억이상": 7, "50억이상": 6, "10억이상": 5, "10억미만": 4}

# 등급별 경력년수 추가 오프셋 (등급이 낮을수록 더 많은 경력 필요)
GRADE_OFFSET = {"특급": 0, "고급": 1, "중급": 2}

# 각 항목의 점수 배분 (최고→최저, 5단계)
# 전기분야: 등급 무관(공사비만), 참여분야/보조/비상주: 등급+공사비
CAREER_SCORES = {
    "책임_전기":    [10.0, 9.0, 8.0, 7.0, 6.0],    # 공사비만, 등급 무관
    "책임_참여1":   [10.0, 8.5, 7.0, 5.5, 4.0],    # 공사비+등급
    "책임_참여2":   [5.0, 4.5, 4.0, 3.5, 3.0],     # 공사비+등급 (감독/감리만)
    "보조_전기":    [6.0, 5.5, 5.0, 4.5, 4.0],     # 공사비만, 등급 무관
    "보조_참여1":   [6.0, 5.2, 4.4, 3.6],           # 고정기준(5년), 공사비/등급 무관
    "보조_참여2":   [3.0, 2.8, 2.6, 2.4],           # 고정기준(5년), 공사비/등급 무관
    "비상주_전기":  [7.0, 6.5, 6.0, 5.5, 5.0],     # 공사비만, 등급 무관
}

# 등급이 없는(공사비만 적용) 항목 목록
NO_GRADE_ITEMS = {"책임_전기", "보조_전기", "비상주_전기"}

# 공사비 구간과 무관하게 고정 base 사용하는 항목 (참여분야 고정 년수기준)
FIXED_BASE_ITEMS = {"보조_참여1": 5, "보조_참여2": 5}


def get_cost_base(cost_tier=None):
    """총 예정전기공사비 구간 → 기본 임계값 반환"""
    if cost_tier is None:
        cost_tier = COST_TIER
    return COST_TIER_BASE.get(cost_tier, 4)


def calc_career_score(years, item_key, grade="", cost_tier=None):
    """경력년수로 점수 계산 (범용)

    Args:
        years: 경력 년수 (float, 1년미만 절사)
        item_key: CAREER_SCORES 키 (예: "책임_전기", "책임_참여1")
        grade: 등급 ("특급"/"고급"/"중급"), 등급무관 항목은 무시됨
        cost_tier: 공사비 구간 (None이면 전역 COST_TIER 사용)

    Returns:
        (점수, 근거문자열)
    """
    scores = CAREER_SCORES.get(item_key)
    if not scores:
        return 0.0, ""

    y = int(years)  # 1년미만 절사

    # 고정 base 항목은 공사비/등급 무관
    if item_key in FIXED_BASE_ITEMS:
        base = FIXED_BASE_ITEMS[item_key]
    else:
        base = get_cost_base(cost_tier)
        # 등급 오프셋 적용 (전기분야는 등급 무관)
        if item_key not in NO_GRADE_ITEMS:
            base += GRADE_OFFSET.get(grade, 2)

    show_grade = grade and item_key not in NO_GRADE_ITEMS and item_key not in FIXED_BASE_ITEMS
    prefix = f"{grade} " if show_grade else ""

    for i, score in enumerate(scores):
        if y >= base - i:
            return score, f"{prefix}{years:.1f}년→{score}점"

    return scores[-1], f"{prefix}{years:.1f}년→{scores[-1]}점"

# 자본비율 점수표
CAPITAL_RATIO_SCORE = [
    (100, 1.0),   # 100% 이상
    (75, 0.85),   # 75% 이상
    (50, 0.7),    # 50% 이상
    (25, 0.55),   # 25% 이상
    (0, 0.4),     # 25% 미만
]

# 유동비율 점수표
CURRENT_RATIO_SCORE = [
    (100, 2.0),   # 100% 이상
    (75, 1.7),    # 75% 이상
    (50, 1.4),    # 50% 이상
    (25, 1.1),    # 25% 이상
    (0, 0.8),     # 25% 미만
]

# 비상주 중첩도 점수표 (6개소미만=4, 6=3, 7=2, 8=1, 9이상=0)
NONRESIDENT_OVERLAP_SCORE = {
    0: 4, 1: 4, 2: 4, 3: 4, 4: 4, 5: 4,
    6: 3, 7: 2, 8: 1, 9: 0,
}

# 가점 - 책임감리원 자격보유
CERT_BONUS = {
    "기술사": 1.0,
    "기능장": 0.8,
    "기사": 0.7,
    "산업기사": 0.5,
}

# 부실벌점 감점표 (누계평균부실벌점 → 감점, 공고일 2개월전 최근2년)
PENALTY_DEDUCTION_SCORE = [
    (20, -5.0),   # 20점 이상
    (15, -3.0),   # 15~20점 미만
    (10, -2.0),   # 10~15점 미만
    (5,  -1.0),   # 5~10점 미만
    (2,  -0.5),   # 2~5점 미만
    (1,  -0.2),   # 1~2점 미만
    (0,   0.0),   # 1점 미만
]

# 유사용역 점수표 (사정금액 대비 비율 → 점수)
SIMILAR_PROJECT_SCORE = [
    (100, 10.0),   # 100% 이상
    (80,  9.0),    # 80% 이상
    (60,  8.0),    # 60% 이상
    (40,  6.0),    # 40% 이상
    (0,   4.0),    # 40% 미만
]

# 비상주감리원 등급 점수 (특급=3, 고급=2, 고급 미만=실격)
NONRESIDENT_GRADE_SCORE = {
    "특급": 3.0,
    "고급": 2.0,
}

# 기술개발 투자실적 점수표 (A÷B, %, 높은 임계값 순 정렬)
TECH_INVEST_SCORE = [
    (3.0, 4.0),    # 3% 이상
    (2.5, 3.5),    # 2.5% 이상
    (2.0, 3.0),    # 2% 이상
    (1.5, 2.5),    # 1.5% 이상
    (0,   2.0),    # 1.5% 미만
]

# 개발실적 점수표 (종류별 × 기간비율%)
# 기간비율은 업체가 기재한 값(100/80/60)을 읽어 해당 타입 기준점수에 적용
TECH_DEV_SCORE = {
    '특허':       {100: 1.00, 80: 0.80, 60: 0.60, 0: 0.0},
    '실용신안':   {100: 0.50, 80: 0.40,            0: 0.0},
    '전력신기술': {100: 2.00, 80: 1.60,            0: 0.0},
}

# 교체빈도 감리업체 점수표 (교체율 %, 높은 임계값 순 정렬)
REPLACEMENT_RATE_SCORE = [
    (50,  1.0),    # 50% 이상
    (30,  1.5),    # 30% 이상 50% 미만
    (10,  2.0),    # 10% 이상 30% 미만
    (0,   2.5),    # 10% 미만
]


###############################################################################
# OCR 엔진 초기화 + 파일 캐시 (JSON 영구 저장)
###############################################################################
import gc
import hashlib

_reader = None
_ocr_memory_cache = {}  # 메모리 캐시 (세션 내 중복 방지)
_current_pdf_path = None
_cache_dir = None


def get_reader():
    global _reader
    if _reader is None:
        print("[INFO] OCR 엔진 초기화 중 (최초 1회)...")
        try:
            _reader = easyocr.Reader(['ko', 'en'], gpu=False)
            print("[INFO] OCR 엔진 준비 완료")
        except Exception as e:
            print(f"[WARN] OCR 엔진 초기화 실패 ({type(e).__name__}): {e}")
            _reader = False  # 실패 표시 (None과 구분)
    if _reader is False:
        return None
    return _reader


def set_current_pdf(pdf_path):
    """현재 처리중인 PDF 설정 → 파일 캐시 디렉토리 결정"""
    global _current_pdf_path, _cache_dir
    _current_pdf_path = pdf_path
    _cache_dir = None


def _get_cache_dir():
    """PDF별 캐시 디렉토리 (증빙/.ocr_cache/<파일명>_<크기>/)"""
    global _cache_dir
    if _cache_dir is not None:
        return _cache_dir
    if not _current_pdf_path:
        return None
    base_dir = os.path.join(os.path.dirname(_current_pdf_path), ".ocr_cache")
    pdf_name = os.path.splitext(os.path.basename(_current_pdf_path))[0]
    pdf_size = os.path.getsize(_current_pdf_path)
    _cache_dir = os.path.join(base_dir, f"{pdf_name}_{pdf_size}")
    os.makedirs(_cache_dir, exist_ok=True)
    return _cache_dir


def _load_page_cache(page_num):
    """파일 캐시에서 페이지 OCR 결과 로드"""
    cache_dir = _get_cache_dir()
    if not cache_dir:
        return None
    cache_file = os.path.join(cache_dir, f"page_{page_num:04d}.json")
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [(item["text"], item["conf"], item["bbox"]) for item in data]
        except (json.JSONDecodeError, KeyError, TypeError):
            return None
    return None


def _to_native(obj):
    """numpy 타입을 Python 네이티브 타입으로 변환"""
    if isinstance(obj, (list, tuple)):
        return [_to_native(x) for x in obj]
    try:
        return obj.item()  # numpy scalar → Python int/float
    except (AttributeError, ValueError):
        return obj


def _save_page_cache(page_num, parsed):
    """OCR 결과를 JSON 파일로 영구 저장"""
    cache_dir = _get_cache_dir()
    if not cache_dir:
        return
    cache_file = os.path.join(cache_dir, f"page_{page_num:04d}.json")
    data = [{"text": t, "conf": float(c), "bbox": _to_native(b)} for (t, c, b) in parsed]
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
    except Exception as e:
        print(f"  [WARN] 캐시 저장 실패 (p{page_num+1}): {e}")


def ocr_page(doc, page_num, zoom=1.0):
    """PDF 페이지를 OCR하여 텍스트 리스트 반환 (메모리+파일 캐시)"""
    # 1) 메모리 캐시
    cache_key = (id(doc), page_num)
    if cache_key in _ocr_memory_cache:
        return _ocr_memory_cache[cache_key]

    # 2) 파일(JSON) 캐시 → 있으면 OCR 건너뛰기
    cached = _load_page_cache(page_num)
    if cached is not None:
        _ocr_memory_cache[cache_key] = cached
        return cached

    # 3) OCR 실행
    reader = get_reader()
    page = doc[page_num]
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat)
    img_data = pix.tobytes('png')
    del pix
    gc.collect()

    try:
        results = reader.readtext(img_data, batch_size=1)
    except Exception as e:
        # numpy ArrayMemoryError, MemoryError, RuntimeError 등 모두 캐치
        print(f"  [WARN] 페이지 {page_num+1} OCR 실패({type(e).__name__}), 저해상도 재시도...")
        del img_data
        gc.collect()
        # 0.8배로 축소하여 재시도
        mat = fitz.Matrix(0.8, 0.8)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes('png')
        del pix
        gc.collect()
        try:
            results = reader.readtext(img_data, batch_size=1)
        except Exception as e2:
            print(f"  [ERROR] 페이지 {page_num+1} OCR 완전 실패: {type(e2).__name__}")
            del img_data
            gc.collect()
            return []

    parsed = [(text, conf, bbox) for (bbox, text, conf) in results]

    # 캐시 저장 (메모리 + 파일)
    _ocr_memory_cache[cache_key] = parsed
    _save_page_cache(page_num, parsed)

    del img_data, results
    gc.collect()
    return parsed


def ocr_page_text(doc, page_num, zoom=1.0):
    """페이지의 전체 텍스트를 하나의 문자열로 반환"""
    results = ocr_page(doc, page_num, zoom)
    return ' '.join([text for (text, conf, bbox) in results])


###############################################################################
# 텍스트 추출: PyMuPDF 우선 → OCR 폴백
###############################################################################

def get_page_text(doc, page_num):
    """페이지 텍스트 추출 (PyMuPDF 내장 텍스트 우선, 부족하면 OCR 폴백)"""
    if page_num >= doc.page_count:
        return ""
    page = doc[page_num]
    text = page.get_text()
    if text and len(text.strip()) > 20:
        return text
    # 텍스트가 없거나 부족하면 OCR 폴백
    return ocr_page_text(doc, page_num)


def get_page_lines(doc, page_num):
    """페이지의 줄 단위 텍스트 리스트 반환"""
    text = get_page_text(doc, page_num)
    return [l.strip() for l in text.split('\n') if l.strip()]


###############################################################################
# 데이터 추출 함수 (PyMuPDF 텍스트 우선 방식)
###############################################################################

def smart_find_page(doc, keywords, search_range, label=""):
    """핵심 페이지를 빠르게 찾기 - PyMuPDF 내장 텍스트만 사용 (OCR 없음, 크래시 방지)"""
    for page_num in search_range:
        if page_num >= doc.page_count:
            break
        text = doc[page_num].get_text()  # OCR 폴백 없이 내장 텍스트만
        if any(kw in text for kw in keywords):
            if label:
                print(f"    [{label}] 페이지 {page_num+1}에서 발견")
            return page_num
    return -1


def classify_pages(doc):
    """전체 페이지를 스캔하여 양식별로 분류 → 페이지 맵 반환

    Returns: {"양식2-4": [16], "양식2-5_책임": [20], ...}
    """
    page_map = {}
    # 우선순위 순서: 고유 식별자가 있는 양식 먼저 매칭
    FORM_RULES = [
        ("양식2-3", [["종합득점표"], ["양식2-3"]]),
        ("양식2-4", [["양식2-4"], ["참여감리원", "자격사항", "자격등급"]]),
        ("양식2-5_책임", [["양식2-5", "1. 책임감리원"], ["양식2-5", "책임감리원"]]),
        ("양식2-5_보조", [["양식2-5", "2. 보조감리원"], ["양식2-5", "보조감리원"]]),
        ("양식2-5_비상주", [["3. 비상주감리원"], ["양식2-5", "비상주감리원"]]),
        ("양식2-6", [["유사용역", "환산금액"], ["양식2-6"]]),
        ("양식2-9", [["양식2-9"], ["기술개발투자실적"]]),
        ("양식2-10", [["양식2-10"], ["업무중첩도", "배치현황"]]),
        ("양식2-11", [["양식2-11", "교체빈도"]]),
    ]
    for pn in range(doc.page_count):
        text = doc[pn].get_text()
        if not text.strip():
            continue
        for form_id, keyword_sets in FORM_RULES:
            if any(all(kw in text for kw in kws) for kws in keyword_sets):
                page_map.setdefault(form_id, []).append(pn)
                break
    return page_map


# ── find_tables 헬퍼 함수 ──

_NAME_FILTER = {'책임감리원', '보조감리원', '비상주', '감리원', '전기', '토목',
                '특급', '고급', '중급', '초급', '소계', '합계', '배점', '평점',
                '자격사항', '자격내용', '자격등급', '참여감리원', '학력내용',
                '소속회사', '생년월일', '비상주감리원', '전기분야', '참여분야',
                '교육수료', '종료일자', '최종학력', '취득일', '자격명'}

_NAME_EXCLUDE_KW = ['감리', '기술', '분야', '비고', '양식', '교육', '재직', '증명',
                    '사본', '추가', '확인', '발급', '등본', '경우', '첨부', '상주',
                    '법인', '협회', '시행', '건설', '경력', '학위', '자격', '평가',
                    '공사', '용역', '시행']


def _extract_name_from_cells(row):
    """테이블 행에서 한글 이름(2~4자) 추출 — 공백 포함 이름 대응 ('윤 민' 등)"""
    for cell in row:
        if cell is None:
            continue
        val = cell.strip()
        # 1) 연속 한글 2~4자
        m = re.search(r'(?<![가-힣])([가-힣]{2,4})(?![가-힣])', val)
        if m:
            name = m.group(1)
            if name not in _NAME_FILTER and not any(kw in name for kw in _NAME_EXCLUDE_KW):
                return name
        # 2) 공백 포함 한글 이름 ('윤 민', '이 수 진' 등)
        m2 = re.search(r'(?<![가-힣])([가-힣]\s+[가-힣]{1,3})(?![가-힣])', val)
        if m2:
            name = m2.group(1).replace(' ', '')
            if len(name) >= 2 and name not in _NAME_FILTER and not any(kw in name for kw in _NAME_EXCLUDE_KW):
                return name
    return ""


def _extract_grade_from_cells(row):
    """테이블 행에서 등급(특급/고급/중급/초급) 추출"""
    for cell in row:
        if cell is None:
            continue
        for g in ('특급', '고급', '중급', '초급'):
            if g in cell:
                return g
    return ""


def _extract_date_from_cells(row):
    """테이블 행에서 생년월일(YYYY-MM-DD) 추출"""
    for cell in row:
        if cell is None:
            continue
        m = re.search(r'(\d{4}-\d{2}-\d{2})', cell)
        if m:
            year = int(m.group(1)[:4])
            if 1940 < year < 2000:
                return m.group(1)
    return ""


def _extract_cert_from_cells(row):
    """테이블 행에서 자격증명 추출"""
    for cell in row:
        if cell is None:
            continue
        m = re.search(r'(전기[가-힣]*기[사술장]|[가-힣]*기[사술장])', cell)
        if m:
            return m.group(1)
    return ""


def extract_personnel_v2(doc, page_map):
    """양식2-4 find_tables() 기반 인력 추출 — 업체별 양식 차이 자동 대응

    Variant A (성문): 역할 행 + 다음 행에 이름 (cell[0]=None, cell[1]=이름)
    Variant B (로운/다은): 역할 + 이름 같은 행 (cell[0]=역할, cell[1]=이름)

    Returns: personnel dict 또는 None (find_tables 실패 시)
    """
    personnel = {
        "책임감리원": {"성명": "", "등급": "", "자격증": "", "생년월일": ""},
        "보조감리원": [],
        "비상주감리원": {"성명": "", "등급": "", "자격증": "", "생년월일": ""},
        "page": -1,
    }

    pages = page_map.get("양식2-4", [])
    if not pages:
        return None

    page_num = pages[0]
    personnel["page"] = page_num + 1
    page = doc[page_num]

    tables = page.find_tables()
    if not tables.tables:
        return None

    rows = tables.tables[0].extract()
    if len(rows) < 4 or len(rows[0]) < 10:
        return None

    current_role = None
    bozo_person = None  # 현재 보조감리원 dict 참조

    for i in range(2, len(rows)):  # 헤더 2행 skip
        row = rows[i]
        cell0 = (row[0] or '').strip().replace('\n', ' ')

        # 역할 감지
        new_role = None
        if '책임감리원' in cell0:
            new_role = '책임'
        elif '보조감리원' in cell0 or cell0 == '보조':
            new_role = '보조'
        elif '비상주' in cell0:
            new_role = '비상주'

        if new_role:
            current_role = new_role
            if new_role == '보조':
                bozo_person = {"성명": "", "등급": "", "경력일수": ""}
                personnel["보조감리원"].append(bozo_person)

        if current_role is None:
            continue

        # 행에서 데이터 추출
        name = _extract_name_from_cells(row)
        grade = _extract_grade_from_cells(row)
        birth = _extract_date_from_cells(row)
        cert = _extract_cert_from_cells(row)

        if current_role == '책임':
            p = personnel["책임감리원"]
            if name and not p["성명"]:
                p["성명"] = name
            if grade and not p["등급"]:
                p["등급"] = grade
            if birth and not p["생년월일"]:
                p["생년월일"] = birth
            if cert and not p["자격증"]:
                p["자격증"] = cert
        elif current_role == '보조' and bozo_person:
            if name and not bozo_person["성명"]:
                bozo_person["성명"] = name
            if grade and not bozo_person["등급"]:
                bozo_person["등급"] = grade
        elif current_role == '비상주':
            p = personnel["비상주감리원"]
            if name and not p["성명"]:
                p["성명"] = name
            if grade and not p["등급"]:
                p["등급"] = grade
            if birth and not p["생년월일"]:
                p["생년월일"] = birth
            if cert and not p["자격증"]:
                p["자격증"] = cert

    # 검증: 최소한 책임감리원 이름+등급이 있어야 유효
    if not personnel["책임감리원"]["성명"] or not personnel["책임감리원"]["등급"]:
        return None

    print(f"    [v2] 책임: {personnel['책임감리원']['성명']}/{personnel['책임감리원']['등급']}, "
          f"보조: {personnel['보조감리원'][0]['성명'] if personnel['보조감리원'] else '-'}/"
          f"{personnel['보조감리원'][0]['등급'] if personnel['보조감리원'] else '-'}, "
          f"비상주: {personnel['비상주감리원']['성명']}/{personnel['비상주감리원']['등급']}")

    return personnel


def extract_career_v2(doc, page_map, role="책임"):
    """양식2-5 경력 추출 통합 — find_tables() + get_text() 하이브리드

    role: "책임"|"보조"|"비상주"
    Returns: 기존 함수와 동일한 dict 형식 또는 None (실패 시)
    """
    if role == "책임":
        result = {
            "책임_전기분야_개월": 0, "책임_전기분야_년": 0,
            "책임_참여분야1_개월": 0, "책임_참여분야1_년": 0,
            "책임_참여분야2_개월": 0, "책임_참여분야2_년": 0,
            "책임_전기분야_점수": 0, "책임_참여분야1_점수": 0, "책임_참여분야2_점수": 0,
            "page": -1,
        }
        page_key = "양식2-5_책임"
        prefix = "책임"
        score_limits = {"전기": 10, "참여1": 10, "참여2": 5}
    elif role == "보조":
        result = {
            "보조_전기분야_년": 0, "보조_전기분야_점수": 0,
            "보조_참여분야1_년": 0, "보조_참여분야1_점수": 0,
            "보조_참여분야2_년": 0, "보조_참여분야2_점수": 0,
            "보조_등급": "",
            "page": -1,
        }
        page_key = "양식2-5_보조"
        prefix = "보조"
        score_limits = {"전기": 6, "참여1": 6, "참여2": 3}
    elif role == "비상주":
        result = {
            "비상주_전기분야_년": 0, "비상주_전기분야_점수": 0,
            "비상주_등급_점수": 0,
            "page": -1,
        }
        page_key = "양식2-5_비상주"
        prefix = "비상주"
        score_limits = {"전기": 7}
    else:
        return None

    pages = page_map.get(page_key, [])
    if not pages:
        return None

    # 첫 번째 페이지를 기본으로, 다중 페이지면 계 행이 있는 페이지 찾기
    result["page"] = pages[0] + 1

    # ── 1단계: find_tables() → 경력 개월수 추출 (계 행) ──
    sum_found = False
    for pn in pages:
        page = doc[pn]
        tables = page.find_tables()
        if not tables.tables:
            continue
        for t in tables.tables:
            rows = t.extract()
            for row in rows:
                if (row[0] or '').strip() == '계':
                    ncols = len(row)
                    if role in ("책임", "보조") and ncols >= 10:
                        try:
                            elec = float((row[4] or '0').replace(',', ''))
                            result[f"{prefix}_전기분야_년"] = round(elec / 12, 2)
                            if role == "책임":
                                result["책임_전기분야_개월"] = elec
                        except (ValueError, TypeError):
                            pass
                        try:
                            p1 = float((row[6] or '0').replace(',', ''))
                            result[f"{prefix}_참여분야1_년"] = round(p1 / 12, 2)
                            if role == "책임":
                                result["책임_참여분야1_개월"] = p1
                        except (ValueError, TypeError):
                            pass
                        try:
                            p2 = float((row[9] or '0').replace(',', ''))
                            result[f"{prefix}_참여분야2_년"] = round(p2 / 12, 2)
                            if role == "책임":
                                result["책임_참여분야2_개월"] = p2
                        except (ValueError, TypeError):
                            pass
                        sum_found = True
                    elif role == "비상주" and ncols >= 7:
                        try:
                            elec = float((row[6] or '0').replace(',', ''))
                            result["비상주_전기분야_년"] = round(elec / 12, 2)
                        except (ValueError, TypeError):
                            pass
                        sum_found = True
                    break
            if sum_found:
                break
        if sum_found:
            break

    # ── 2단계: get_text() → 점수/등급 추출 (첫 번째 페이지 상단) ──
    text = doc[pages[0]].get_text()
    lines = [l.strip() for l in text.split('\n')]

    if role in ("책임", "보조"):
        # 점수: 라. 전기분야 / 마-1. 참여분야 / 마-2. 참여분야
        score_section = None
        for i, line in enumerate(lines):
            if '라' in line and '전기분야' in line:
                score_section = "전기"
                continue
            elif '마-2' in line:
                score_section = "참여2"
                continue
            elif '마-1' in line or ('마' in line and '참여분야' in line and '마-2' not in line):
                score_section = "참여1"
                continue

            if score_section is None:
                continue

            score = None
            m = re.search(r'(\d+\.?\d*)\s*점', line)
            if m:
                score = float(m.group(1))
            elif re.match(r'^\d+\.?\d*$', line):
                # 숫자만 있는 줄 → 다음 비어있지 않은 줄에 '점'이 있는지 확인
                for j in range(i + 1, min(i + 3, len(lines))):
                    if lines[j]:
                        if '점' in lines[j]:
                            score = float(line)
                        break

            if score is not None:
                max_s = score_limits.get(score_section, 99)
                if score_section == "전기":
                    key = f"{prefix}_전기분야_점수"
                elif score_section == "참여1":
                    key = f"{prefix}_참여분야1_점수"
                else:
                    key = f"{prefix}_참여분야2_점수"

                if result.get(key, 0) == 0 and score <= max_s:
                    result[key] = score

        # 보조 등급 추출
        if role == "보조":
            for line in lines:
                if line in ('특급', '고급', '중급', '초급'):
                    result["보조_등급"] = line
                    break

    elif role == "비상주":
        # 등급 점수: 특급→3, 고급→2
        for line in lines:
            if line in ('특급', '고급'):
                result["비상주_등급_점수"] = 3.0 if line == '특급' else 2.0
                break

        # 전기분야 점수: "다. 전기분야" 또는 "나. 전기분야" 아래
        score_section = None
        for i, line in enumerate(lines):
            if ('다' in line or '나' in line) and '전기분야' in line:
                score_section = "전기"
                continue
            if score_section == "전기":
                m = re.search(r'(\d+\.?\d*)\s*점', line)
                if m:
                    score = float(m.group(1))
                    if score <= 7:
                        result["비상주_전기분야_점수"] = score
                        break
                elif re.match(r'^\d+\.?\d*$', line):
                    # 숫자만 있는 줄 → 다음 비어있지 않은 줄에 '점'이 있는지 확인
                    maybe_score = float(line)
                    for j in range(i + 1, min(i + 3, len(lines))):
                        if lines[j]:
                            if '점' in lines[j] and maybe_score <= 7:
                                result["비상주_전기분야_점수"] = maybe_score
                            break
                    if result["비상주_전기분야_점수"] > 0:
                        break

    if not sum_found:
        return None  # 테이블에서 계 행을 못 찾으면 실패

    print(f"    [v2-{role}] p{result['page']}: ", end="")
    if role in ("책임", "보조"):
        print(f"전기={result[f'{prefix}_전기분야_년']}년/{result[f'{prefix}_전기분야_점수']}점, "
              f"참여1={result[f'{prefix}_참여분야1_점수']}점, 참여2={result[f'{prefix}_참여분야2_점수']}점")
    elif role == "비상주":
        print(f"등급={result['비상주_등급_점수']}점, 전기={result['비상주_전기분야_년']}년/{result['비상주_전기분야_점수']}점")

    return result


def extract_company_name(doc):
    """업체명 추출 (PyMuPDF 텍스트 우선)"""
    for page_num in range(min(5, doc.page_count)):
        lines = get_page_lines(doc, page_num)
        for line in lines:
            # "(주)XXX" 또는 "㈜XXX" 패턴
            m = re.search(r'[(\(]주[)\)]?\s*[가-힣]+|㈜\s*[가-힣]+|주식회사\s*[가-힣]+', line)
            if m:
                return m.group(0).strip()
    return "미확인"


def extract_summary_table(doc):
    """종합득점표(양식2-3) 추출 - 대괄호 배점 순서 기반"""
    result = {
        "참여감리원": 0,    # [50]
        "유사용역": 0,      # 1st [10]
        "신용도": 0,        # 2nd [10]
        "기술개발": 0,      # 3rd [10]
        "업무중첩도": 0,    # 4th [10]
        "교체빈도": 0,      # 1st [5]
        "작업계획": 0,      # 2nd [5]
        "가감점": 0,        # [  ]
        "총점": 0,
        "page": -1,
    }

    page_num = smart_find_page(doc, ['종합득점표', '양식2-3'],
                                range(10, min(25, doc.page_count)), "종합득점표")
    if page_num < 0:
        page_num = smart_find_page(doc, ['종합득점표'],
                                    range(0, min(30, doc.page_count)), "종합득점표(확장)")

    if page_num >= 0:
        result["page"] = page_num + 1
        lines = get_page_lines(doc, page_num)

        # 대괄호 [XX] 패턴으로 배점 순서 추출
        # 종합득점표 구조: [50] → [10] → [10] → [10] → [10] → [5] → [5] → [ ] → 100(총배점)
        # 각 대괄호 다음 줄의 숫자가 업체제출 점수
        bracket_scores = []  # (배점, 업체제출점수) 쌍
        for i, line in enumerate(lines):
            m = re.match(r'^\s*\[(\d+|[\s]*)\]\s*$', line.strip())
            if m:
                bracket_val = m.group(1).strip()
                # 다음 줄에서 업체제출 점수 찾기
                for j in range(i+1, min(i+3, len(lines))):
                    score_m = re.match(r'^\s*(\d+\.?\d*)\s*$', lines[j].strip())
                    if score_m:
                        score = float(score_m.group(1))
                        bracket_scores.append((bracket_val, score))
                        break

        # 순서대로 매핑: [50], [10]x4, [5]x2, [ ]
        # 항목 이름 순서 (종합득점표 표준)
        category_order = ["참여감리원", "유사용역", "신용도", "기술개발",
                          "업무중첩도", "교체빈도", "작업계획", "가감점"]
        for idx, (bv, score) in enumerate(bracket_scores):
            if idx < len(category_order):
                result[category_order[idx]] = score

        # 총점: "100" 다음 줄
        for i, line in enumerate(lines):
            if line.strip() == '100':
                for j in range(i+1, min(i+3, len(lines))):
                    m = re.match(r'^\s*(\d+\.?\d+)\s*$', lines[j].strip())
                    if m:
                        val = float(m.group(1))
                        if 50 <= val <= 120:
                            result["총점"] = val
                            break

    return result


def extract_personnel(doc):
    """참여감리원 자격사항(양식2-4) 추출 - PyMuPDF 텍스트 사용"""
    personnel = {
        "책임감리원": {"성명": "", "등급": "", "자격증": "", "생년월일": ""},
        "보조감리원": [],
        "비상주감리원": {"성명": "", "등급": "", "자격증": "", "생년월일": ""},
        "page": -1,
    }

    # 양식2-4 페이지 찾기 (종합득점표 양식2-3과 구분 필요)
    page_num = smart_find_page(doc, ['양식2-4'],
                                range(12, min(25, doc.page_count)), "참여감리원")
    if page_num < 0:
        page_num = smart_find_page(doc, ['참여감리원 자격사항', '자격등급'],
                                    range(12, min(30, doc.page_count)), "참여감리원(확장)")

    if page_num >= 0:
        personnel["page"] = page_num + 1
        lines = get_page_lines(doc, page_num)

        grades = ['특급', '고급', '중급', '초급']
        certs = ['기술사', '기능장', '기사', '산업기사']
        # 이름 필터 목록
        name_filter = {'책임감리원', '보조감리원', '비상주', '감리원', '전기', '토목',
                       '특급', '고급', '중급', '초급', '소계', '합계', '배점', '평점',
                       '자격사항', '자격내용', '자격등급', '참여감리원', '학력내용',
                       '소속회사', '생년월일', '비상주감리원', '전기분야', '참여분야',
                       '교육수료', '종료일자', '최종학력', '취득일', '자격명',
                       '성문기술단', '졸업일'}

        # 1단계: 모든 이름 후보 수집 (순서대로)
        all_names = []
        for i, line in enumerate(lines):
            name_match = re.search(r'(?<![가-힣])([가-힣]{2,4})(?![가-힣])', line)
            if name_match:
                name = name_match.group(1)
                if name not in name_filter and not any(kw in name for kw in
                    ['감리', '기술', '분야', '비고', '양식', '교육', '재직', '증명',
                     '사본', '추가', '확인', '발급', '등본', '경우', '첨부', '상주',
                     '법인', '협회', '시행', '건설', '경력', '학위', '자격']):
                    all_names.append((i, name))

        # 2단계: 섹션 마커 위치 파악
        chief_line = -1
        asst_line = -1
        nonres_line = -1
        for i, line in enumerate(lines):
            if '책임감리원' in line and chief_line < 0:
                chief_line = i
            if ('보조감리원' in line or '보조 감리원' in line) and asst_line < 0:
                asst_line = i
            if '비상주' in line and nonres_line < 0:
                nonres_line = i

        # 3단계: 이름 할당 (테이블 구조 - 책임/비상주가 같은 행에 나란히)
        if len(all_names) >= 1:
            personnel["책임감리원"]["성명"] = all_names[0][1]
        if len(all_names) >= 2:
            if asst_line < 0 or all_names[1][0] < asst_line:
                personnel["비상주감리원"]["성명"] = all_names[1][1]
            else:
                personnel["보조감리원"].append({"성명": all_names[1][1], "등급": "", "경력일수": ""})
        for idx, (line_num, name) in enumerate(all_names[2:], 2):
            if asst_line >= 0 and line_num >= asst_line:
                personnel["보조감리원"].append({"성명": name, "등급": "", "경력일수": ""})

        # 4단계: 등급 - 섹션 마커 근접도 기반 할당
        all_grades = []
        for i, line in enumerate(lines):
            for grade in grades:
                if grade == line.strip():
                    all_grades.append((i, grade))

        # 각 등급을 가장 가까운 섹션 마커에 할당
        markers = []
        if chief_line >= 0:
            markers.append(("책임", chief_line))
        if asst_line >= 0:
            markers.append(("보조", asst_line))
        if nonres_line >= 0:
            markers.append(("비상주", nonres_line))

        assigned_grades = {}
        used = set()
        # 각 마커에 대해 가장 가까운 미할당 등급 찾기
        for section, marker_pos in sorted(markers, key=lambda x: x[1]):
            best_idx = -1
            best_dist = float('inf')
            for gi, (gpos, gval) in enumerate(all_grades):
                if gi in used:
                    continue
                dist = abs(gpos - marker_pos)
                if dist < best_dist:
                    best_dist = dist
                    best_idx = gi
            if best_idx >= 0:
                assigned_grades[section] = all_grades[best_idx][1]
                used.add(best_idx)

        if "책임" in assigned_grades:
            personnel["책임감리원"]["등급"] = assigned_grades["책임"]
        if "비상주" in assigned_grades:
            personnel["비상주감리원"]["등급"] = assigned_grades["비상주"]
        if "보조" in assigned_grades and personnel["보조감리원"]:
            personnel["보조감리원"][0]["등급"] = assigned_grades["보조"]

        # 5단계: 자격증/생년월일 추출
        for i, line in enumerate(lines):
            for cert in certs:
                if cert in line:
                    cert_match = re.search(r'(전기[가-힣]*기[사술장]|[가-힣]*기[사술장])', line)
                    if cert_match and not personnel["책임감리원"]["자격증"]:
                        personnel["책임감리원"]["자격증"] = cert_match.group(1)

            birth_match = re.search(r'(\d{4}-\d{2}-\d{2})', line)
            if birth_match:
                bdate = birth_match.group(1)
                year = int(bdate[:4])
                if 1940 < year < 2000:
                    if not personnel["책임감리원"]["생년월일"]:
                        personnel["책임감리원"]["생년월일"] = bdate
                    elif not personnel["비상주감리원"]["생년월일"]:
                        personnel["비상주감리원"]["생년월일"] = bdate

    return personnel


def extract_career_summary(doc):
    """책임감리원 경력실적사항(양식2-5) 추출 - PyMuPDF 텍스트 사용"""
    career = {
        "책임_전기분야_개월": 0,
        "책임_전기분야_년": 0,
        "책임_참여분야1_개월": 0,
        "책임_참여분야1_년": 0,
        "책임_참여분야2_개월": 0,
        "책임_참여분야2_년": 0,
        "책임_전기분야_점수": 0,
        "책임_참여분야1_점수": 0,
        "책임_참여분야2_점수": 0,
        "page": -1,
    }

    page_num = smart_find_page(doc, ['양식2-5', '경력실적사항'],
                                range(17, min(30, doc.page_count)), "경력실적")
    if page_num < 0:
        page_num = smart_find_page(doc, ['책임감리원', '전기분야'],
                                    range(15, min(35, doc.page_count)), "경력실적(확장)")

    if page_num >= 0:
        career["page"] = page_num + 1
        lines = get_page_lines(doc, page_num)

        section = None  # "전기", "참여1", "참여2"
        months_count = 0
        score_count = 0

        for line in lines:
            # 섹션 판별
            if '가.' in line and '전기분야' in line:
                section = "전기"
            elif '나-1' in line or ('나.' in line and '참여분야' in line and '감독' not in line and '감리' not in line):
                section = "참여1"
            elif '나-2' in line or ('참여분야' in line and ('감독' in line or '감리' in line)):
                section = "참여2"

            # 개월수 추출 (예: 409.33개월)
            m = re.search(r'(\d+\.?\d*)\s*개월', line)
            if m:
                months = float(m.group(1))
                if section == "전기" and career["책임_전기분야_개월"] == 0:
                    career["책임_전기분야_개월"] = months
                    career["책임_전기분야_년"] = round(months / 12, 2)
                elif section == "참여1" and career["책임_참여분야1_개월"] == 0:
                    career["책임_참여분야1_개월"] = months
                    career["책임_참여분야1_년"] = round(months / 12, 2)
                elif section == "참여2" and career["책임_참여분야2_개월"] == 0:
                    career["책임_참여분야2_개월"] = months
                    career["책임_참여분야2_년"] = round(months / 12, 2)
                elif career["책임_전기분야_개월"] == 0:
                    career["책임_전기분야_개월"] = months
                    career["책임_전기분야_년"] = round(months / 12, 2)

            # 점수 추출 (예: "10 점", "라. 전기분야 10 점")
            m = re.search(r'(\d+\.?\d*)\s*점', line)
            if m:
                score = float(m.group(1))
                if '라' in line or '전기분야' in line:
                    if score <= 10:
                        career["책임_전기분야_점수"] = score
                elif '마-1' in line or ('참여분야' in line and '감독' not in line):
                    if score <= 10:
                        career["책임_참여분야1_점수"] = score
                elif '마-2' in line or ('참여분야' in line and ('감독' in line or '감리' in line)):
                    if score <= 5:
                        career["책임_참여분야2_점수"] = score
                else:
                    # 순서대로 할당
                    if career["책임_전기분야_점수"] == 0 and score <= 10:
                        career["책임_전기분야_점수"] = score
                    elif career["책임_참여분야1_점수"] == 0 and score <= 10:
                        career["책임_참여분야1_점수"] = score
                    elif career["책임_참여분야2_점수"] == 0 and score <= 5:
                        career["책임_참여분야2_점수"] = score

    return career


def extract_asst_career(doc, chief_page):
    """보조감리원 경력실적사항(양식2-5 보조) 추출"""
    result = {
        "보조_전기분야_년": 0, "보조_전기분야_점수": 0,
        "보조_참여분야1_년": 0, "보조_참여분야1_점수": 0,
        "보조_참여분야2_년": 0, "보조_참여분야2_점수": 0,
        "보조_등급": "",
        "page": -1,
    }

    # 책임감리원 페이지 다음부터 '보조감리원' 포함 양식2-5 탐색
    start = max(chief_page + 1, 17)
    page_num = smart_find_page(doc, ['양식2-5', '보조감리원'],
                                range(start, min(start + 15, doc.page_count)), "보조감리원경력")
    if page_num < 0:
        page_num = smart_find_page(doc, ['보조감리원', '전기분야', '경력'],
                                    range(start, min(start + 15, doc.page_count)), "보조감리원경력(확장)")
    if page_num < 0:
        return result

    result["page"] = page_num + 1
    lines = get_page_lines(doc, page_num)

    # 경력 개월수 추출
    section = None
    for line in lines:
        if '가.' in line and '전기분야' in line:
            section = "전기"
        elif '나-1' in line or ('나.' in line and '참여분야' in line and '감독' not in line and '감리' not in line):
            section = "참여1"
        elif '나-2' in line or ('참여분야' in line and ('감독' in line or '감리' in line)):
            section = "참여2"

        m = re.search(r'(\d+\.?\d*)\s*개월', line)
        if m:
            months = float(m.group(1))
            years = round(months / 12, 2)
            if section == "전기" and result["보조_전기분야_년"] == 0:
                result["보조_전기분야_년"] = years
            elif section == "참여1" and result["보조_참여분야1_년"] == 0:
                result["보조_참여분야1_년"] = years
            elif section == "참여2" and result["보조_참여분야2_년"] == 0:
                result["보조_참여분야2_년"] = years
            elif result["보조_전기분야_년"] == 0:
                result["보조_전기분야_년"] = years

    # 점수 추출 — 섹션 추적 + 점수/점 분리 라인 처리
    score_section = None
    for i, line in enumerate(lines):
        # 섹션 헤더 감지 (마-2를 마-1보다 먼저 체크)
        if '라' in line and '전기분야' in line:
            score_section = "전기"
            continue
        elif '마-2' in line or ('참여분야' in line and ('감독' in line or '감리' in line)):
            score_section = "참여2"
            continue
        elif '마-1' in line or ('마' in line and '참여분야' in line):
            score_section = "참여1"
            continue

        if score_section is None:
            continue

        score = None
        # 같은 라인: "6 점" 또는 "6.0 점"
        m = re.search(r'(\d+\.?\d*)\s*점', line)
        if m:
            score = float(m.group(1))
        # 분리 라인: "6.0" 다음 줄이 "점"
        elif re.match(r'^\d+\.?\d*$', line) and i + 1 < len(lines) and '점' in lines[i + 1]:
            score = float(line)

        if score is not None:
            if score_section == "전기" and result["보조_전기분야_점수"] == 0 and score <= 6:
                result["보조_전기분야_점수"] = score
            elif score_section == "참여1" and result["보조_참여분야1_점수"] == 0 and score <= 6:
                result["보조_참여분야1_점수"] = score
            elif score_section == "참여2" and result["보조_참여분야2_점수"] == 0 and score <= 3:
                result["보조_참여분야2_점수"] = score

    # 등급 추출
    grades = ['특급', '고급', '중급', '초급']
    for line in lines:
        for g in grades:
            if g == line.strip():
                result["보조_등급"] = g
                break

    print(f"  [보조감리원경력] p{page_num+1}: 전기={result['보조_전기분야_점수']}점, "
          f"참여1={result['보조_참여분야1_점수']}점, 참여2={result['보조_참여분야2_점수']}점")
    return result


def extract_nonres_career(doc, chief_page):
    """비상주감리원 경력실적사항(양식2-5 비상주) 추출"""
    result = {
        "비상주_전기분야_년": 0, "비상주_전기분야_점수": 0,
        "비상주_등급_점수": 0,
        "page": -1,
    }

    # 책임감리원 페이지 다음부터 탐색, '비상주감리원' 섹션 헤더가 있는 페이지 찾기
    start = max(chief_page + 1, 17)
    page_num = -1
    for pn in range(start, min(start + 25, doc.page_count)):
        if pn >= doc.page_count:
            break
        text = doc[pn].get_text()
        # '비상주감리원'이 섹션 헤더로 존재하는 페이지 (설명 텍스트 내 '비상주' 제외)
        if '비상주감리원' in text and ('등' in text and '급' in text or '전기분야' in text):
            # 책임/보조 페이지가 아닌지 확인
            if '1. 책임감리원' not in text and '2. 보조감리원' not in text:
                page_num = pn
                print(f"    [비상주감리원경력] 페이지 {pn+1}에서 발견")
                break
    if page_num < 0:
        return result

    result["page"] = page_num + 1
    lines = get_page_lines(doc, page_num)

    # 등급 점수: "특급" → 3점, "고급" → 2점
    for line in lines:
        if line.strip() in ('특급', '고급'):
            result["비상주_등급_점수"] = 3.0 if line.strip() == '특급' else 2.0
            break

    # 전기분야 경력 개월수 (합계 개월)
    for i, line in enumerate(lines):
        m = re.search(r'(\d+\.?\d*)\s*개월', line)
        if m:
            months = float(m.group(1))
            result["비상주_전기분야_년"] = round(months / 12, 2)
            break

    # 전기분야 점수 — 같은 라인 또는 분리 라인 처리
    score_section = None
    for i, line in enumerate(lines):
        if ('다' in line or '나' in line) and '전기분야' in line:
            score_section = "전기"
            continue
        if score_section == "전기":
            m = re.search(r'(\d+\.?\d*)\s*점', line)
            if m:
                score = float(m.group(1))
                if score <= 7:
                    result["비상주_전기분야_점수"] = score
                    break
            elif re.match(r'^\d+\.?\d*$', line) and i + 1 < len(lines) and '점' in lines[i + 1]:
                score = float(line)
                if score <= 7:
                    result["비상주_전기분야_점수"] = score
                    break

    print(f"  [비상주감리원경력] p{page_num+1}: 등급점수={result['비상주_등급_점수']}점, "
          f"전기={result['비상주_전기분야_점수']}점")
    return result


def extract_similar_project(doc):
    """유사용역 수행실적 추출 - 여러 페이지에 걸쳐 합계 찾기"""
    result = {
        "적용금액": 0,
        "사정금액": 0,
        "비율": 0,
        "점수": 0,
        "page": -1,
    }

    # 유사용역 시작 페이지 찾기
    start_page = smart_find_page(doc, ['유사용역실적', '유사용역', '환산금액'],
                                range(20, min(int(doc.page_count * 0.5), doc.page_count)), "유사용역")

    if start_page >= 0:
        result["page"] = start_page + 1

        # 시작 페이지부터 최대 10페이지 범위에서 모든 줄 수집
        all_lines = []
        for page_num in range(start_page, min(start_page + 10, doc.page_count)):
            text = doc[page_num].get_text()
            if not text.strip():
                break  # 빈 페이지 도달 시 중단
            all_lines.extend([l.strip() for l in text.split('\n') if l.strip()])

        # 1. "합계 제출" 키워드 다음 줄들에서 큰 숫자 찾기 (합계 제출값)
        #    PDF 레이아웃상 합계값이 키워드와 다른 줄에 위치하는 구조 처리
        for i, line in enumerate(all_lines):
            if '합계 제출' in line or '합계제출' in line:
                for j in range(i + 1, min(i + 20, len(all_lines))):
                    try:
                        val = int(all_lines[j].replace(',', ''))
                        if val > 100000:
                            result["적용금액"] = val
                            break
                    except ValueError:
                        pass
                break

        # 2. 비율 (사정금액 대비, 예: 912%) - 100% 초과값만
        for line in all_lines:
            m = re.search(r'(\d+)\s*%', line)
            if m:
                val = int(m.group(1))
                if val > 100:
                    result["비율"] = val

        # 3. 평점/점수 추출 (평점 키워드 다음 줄 포함)
        for i, line in enumerate(all_lines):
            if line == '평점':
                for j in range(i + 1, min(i + 3, len(all_lines))):
                    try:
                        score = float(all_lines[j])
                        if 0 < score <= 10:
                            result["점수"] = score
                            break
                    except ValueError:
                        pass
            elif '평점' in line:
                m = re.search(r'(\d+\.?\d*)', line)
                if m:
                    score = float(m.group(1))
                    if 0 < score <= 10:
                        result["점수"] = score

        # 4. 적용금액이 0이면 누계 실적에서 최대값으로 fallback
        #    "누계 실적" 키워드 다음 줄에서 숫자 추출 후 최대값 사용
        if result["적용금액"] == 0:
            for i, line in enumerate(all_lines):
                if '누계 실적' in line or '누계실적' in line:
                    for j in range(i + 1, min(i + 5, len(all_lines))):
                        try:
                            val = int(all_lines[j].replace(',', ''))
                            if val > 100000:
                                result["적용금액"] = max(result["적용금액"], val)
                        except ValueError:
                            pass

    return result


def extract_financial(doc):
    """재정상태 건실도 추출 - PyMuPDF 텍스트 + OCR 결합"""
    result = {
        "자기자본비율": 0,
        "평균자기자본비율": 0,
        "자본비율_대비": 0,
        "자본비율_점수": 0,
        "유동비율": 0,
        "평균유동비율": 0,
        "유동비율_대비": 0,
        "유동비율_점수": 0,
        "page": -1,
    }

    # 재정상태 페이지 찾기 (공사감리용역수행현황확인서)
    page_num = smart_find_page(doc, ['자기자본비율', '유동비율', '수행현황확인서'],
                                range(25, min(int(doc.page_count * 0.5), doc.page_count)), "재정상태")
    if page_num < 0:
        page_num = smart_find_page(doc, ['자기자본', '유동비율'],
                                    range(20, min(int(doc.page_count * 0.6), doc.page_count)), "재정상태(확장)")

    if page_num >= 0:
        result["page"] = page_num + 1

        # PyMuPDF 텍스트 시도
        text = get_page_text(doc, page_num)
        lines = [l.strip() for l in text.split('\n') if l.strip()]

        # 자기자본비율 추출
        for i, line in enumerate(lines):
            if '자기자본비율' in line:
                # 같은 줄 또는 다음 몇 줄에서 숫자 찾기
                search_text = ' '.join(lines[i:min(i+5, len(lines))])
                nums = re.findall(r'(\d+\.?\d+)', search_text)
                nums_float = [float(n) for n in nums if 0 < float(n) < 200]
                if len(nums_float) >= 2:
                    result["자기자본비율"] = nums_float[0]
                    result["평균자기자본비율"] = nums_float[1]
                elif len(nums_float) == 1:
                    result["자기자본비율"] = nums_float[0]

            if '유동비율' in line and '자기자본' not in line:
                search_text = ' '.join(lines[i:min(i+5, len(lines))])
                nums = re.findall(r'(\d+\.?\d+)', search_text)
                nums_float = [float(n) for n in nums if 0 < float(n) < 500]
                if len(nums_float) >= 2:
                    result["유동비율"] = nums_float[0]
                    result["평균유동비율"] = nums_float[1]
                elif len(nums_float) == 1:
                    result["유동비율"] = nums_float[0]

        # PyMuPDF로 못 찾았으면 OCR 시도 (캐시된 결과만 사용, 새 OCR 실행 안함)
        if result["자기자본비율"] == 0:
            try:
                items = ocr_page(doc, page_num)
                if items:
                    texts = [t for (t, c, b) in items]
                    for i, t in enumerate(texts):
                        if '자기자본' in t:
                            found_nums = []
                            for j in range(i+1, min(i+8, len(texts))):
                                m = re.search(r'(\d[\d,]*\.?\d*)', texts[j])
                                if m:
                                    val = float(m.group(1).replace(',', ''))
                                    if 0 < val < 200:
                                        found_nums.append(val)
                            if len(found_nums) >= 2:
                                result["자기자본비율"] = found_nums[0]
                                result["평균자기자본비율"] = found_nums[1]
            except Exception as e:
                print(f"  [WARN] 재정상태 OCR 폴백 실패: {type(e).__name__}")

        # 대비 계산
        if result["평균자기자본비율"] > 0:
            result["자본비율_대비"] = round(result["자기자본비율"] / result["평균자기자본비율"] * 100, 2)
        if result["평균유동비율"] > 0:
            result["유동비율_대비"] = round(result["유동비율"] / result["평균유동비율"] * 100, 2)

        # 점수 계산
        for threshold, score in CAPITAL_RATIO_SCORE:
            if result["자본비율_대비"] >= threshold:
                result["자본비율_점수"] = score
                break
        for threshold, score in CURRENT_RATIO_SCORE:
            if result["유동비율_대비"] >= threshold:
                result["유동비율_점수"] = score
                break

    return result


def extract_tech_development(doc, bidding_date=BIDDING_DATE):
    """기술개발 및 투자실적 추출 (양식2-9, 텍스트 페이지)

    반환 키:
      개발실적_항목수, 개발실적_계산점수, 개발실적_항목
      투자액_합계, 매출액_합계, 투자비율, 기술투자_계산점수
      교육실적_기재점수
      page
    """
    result = {
        "개발실적_항목수": 0,
        "개발실적_계산점수": 0.0,
        "개발실적_항목": [],
        "투자액_합계": 0,
        "매출액_합계": 0,
        "투자비율": 0.0,
        "기술투자_계산점수": 0.0,
        "교육실적_기재점수": 0.0,
        "page": -1,
    }

    # 입찰공고일 파싱
    try:
        bid_dt = datetime.strptime(bidding_date, '%Y-%m-%d')
    except (ValueError, TypeError):
        bid_dt = datetime.now()

    # 페이지 탐색: 양식2-9 또는 기술개발투자실적 키워드
    tech_start = max(int(doc.page_count * 0.5), 40)
    page_num = smart_find_page(doc, ['양식2-9', '기술개발투자실적', '개발실적'],
                                range(tech_start, doc.page_count), "기술개발")
    if page_num < 0:
        return result

    result["page"] = page_num + 1

    # 여러 페이지에 걸쳐 있을 수 있으므로 최대 3페이지 텍스트 합산
    full_text = ""
    for pn in range(page_num, min(page_num + 3, doc.page_count)):
        full_text += get_page_text(doc, pn) + "\n"

    all_lines = [l.strip() for l in full_text.split('\n') if l.strip()]

    # ── 나. 기술개발투자실적 (A/B 합계) ──────────────────────────────
    def _parse_num(s):
        try:
            return int(s.replace(',', '').replace(' ', ''))
        except (ValueError, AttributeError):
            return 0

    in_invest = False
    for i, line in enumerate(all_lines):
        if '기술개발투자실적' in line or ('나.' in line and '투자' in line):
            in_invest = True
        if not in_invest:
            continue

        # (A) 레이블 직전 큰 숫자 = 투자액 합계
        if line == '(A)' or '(A)' == line.strip():
            for j in range(i - 1, max(i - 6, -1), -1):
                v = _parse_num(all_lines[j])
                if v > 1_000_000:
                    result["투자액_합계"] = v
                    break

        # (B) 레이블 직전 큰 숫자 = 매출액 합계
        if line == '(B)' or '(B)' == line.strip():
            for j in range(i - 1, max(i - 6, -1), -1):
                v = _parse_num(all_lines[j])
                if v > 10_000_000:
                    result["매출액_합계"] = v
                    break

    # 투자비율 계산 (A/B × 100 = %)
    if result["매출액_합계"] > 0 and result["투자액_합계"] > 0:
        ratio = result["투자액_합계"] / result["매출액_합계"] * 100
        result["투자비율"] = round(ratio, 4)
        result["기술투자_계산점수"] = lookup_score(ratio, TECH_INVEST_SCORE)

    # ── 가. 개발실적 (타입 역산 + 기간 독립 계산) ──────────────────
    # 형식: 지정일(date) / 유효기간(date) / 비율% / 평점  ← 4줄 1세트
    # 타입은 섹션 헤더가 보통 잘못 매핑되므로 기재평점/비율로 역산
    date_pat = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    score_items = []
    in_dev = False
    i = 0

    while i < len(all_lines):
        line = all_lines[i]

        # 개발실적 섹션 진입
        if '개발실적' in line and not in_dev:
            in_dev = True

        # 기술투자실적 또는 교육실적 시작 → 개발실적 파싱 종료
        if in_dev and ('기술개발투자실적' in line or '교육실적' in line):
            break

        if not in_dev:
            i += 1
            continue

        # 날짜 쌍(지정일 + 유효기간) 감지 → 다음 날짜는 유효기간이므로 건너뜀
        if (date_pat.match(line) and
                i + 1 < len(all_lines) and
                date_pat.match(all_lines[i + 1])):

            reg_date_str = line
            # 경과년수 계산 (지정일 ~ 입찰공고일)
            try:
                reg_dt = datetime.strptime(reg_date_str, '%Y-%m-%d')
                elapsed_years = (bid_dt - reg_dt).days / 365.25
            except ValueError:
                elapsed_years = 0.0

            # 기간 구간 판정
            if elapsed_years <= 0:
                calc_bracket, bracket_label = 0, '미래(무효)'
            elif elapsed_years <= 5:
                calc_bracket, bracket_label = 100, '5년이하'
            elif elapsed_years <= 10:
                calc_bracket, bracket_label = 80, '5~10년'
            elif elapsed_years <= 20:
                calc_bracket, bracket_label = 60, '10~20년'
            else:
                calc_bracket, bracket_label = 0, '20년초과'

            # 기재 비율%와 평점 탐색 (유효기간 라인 다음부터)
            stated_ratio = None
            stated_score = None
            advance_to = i + 2
            for j in range(i + 2, min(i + 7, len(all_lines))):
                tok = all_lines[j]
                if date_pat.match(tok):
                    break  # 다음 아이템 시작 → 중단
                if '%' in tok and stated_ratio is None:
                    try:
                        stated_ratio = float(tok.replace('%', '').strip())
                    except ValueError:
                        pass
                elif stated_ratio is not None and stated_score is None:
                    try:
                        v = float(tok)
                        if 0 < v <= 2.5:
                            stated_score = v
                            advance_to = j + 1
                    except ValueError:
                        pass

            # 타입 역산 (기재평점 / (기재비율/100) = 기준점수)
            item_type = None
            if stated_ratio and stated_score and stated_ratio > 0:
                base = stated_score / (stated_ratio / 100)
                if abs(base - 2.0) < 0.15:
                    item_type = '전력신기술'
                elif abs(base - 1.0) < 0.15:
                    item_type = '특허'
                elif abs(base - 0.5) < 0.1:
                    item_type = '실용신안'

            # 독립 계산: 우리가 계산한 구간 + 역산한 타입으로 재계산
            if item_type and calc_bracket > 0:
                calc_score = _dev_score_from_ratio(item_type, calc_bracket)
            elif item_type:
                calc_score = 0.0  # 20년 초과 또는 미래 → 인정 안 함
            else:
                calc_score = stated_score or 0.0  # 타입 미확인 시 기재값 신뢰

            if stated_ratio is not None:
                score_items.append({
                    "type": item_type or "미확인",
                    "지정일": reg_date_str,
                    "경과년": round(elapsed_years, 1),
                    "기재구간": f"{stated_ratio}%",
                    "계산구간": f"{calc_bracket}%({bracket_label})",
                    "구간일치": (stated_ratio is not None and
                                abs(stated_ratio - calc_bracket) < 1),
                    "기재평점": stated_score,
                    "계산평점": calc_score,
                    "일치": abs((stated_score or 0) - calc_score) < 0.02,
                })

            i = advance_to  # 유효기간+비율+평점 건너뜀
            continue

        i += 1

    raw_calc = sum(item["계산평점"] for item in score_items)
    result["개발실적_계산점수"] = round(min(raw_calc, 4.0), 2)
    result["개발실적_항목수"] = len(score_items)
    result["개발실적_항목"] = score_items

    # ── 다. 교육실적 기재점수 ────────────────────────────────────────
    # 소계 뒤 숫자를 탐색하되, 없으면 섹션 시작 후 첫 합계 숫자(>1)를 사용
    in_edu = False
    for i, line in enumerate(all_lines):
        if '교육실적' in line:
            in_edu = True
        if not in_edu:
            continue
        # 소계 직후 숫자
        if '소 계' in line or '소계' in line:
            for j in range(i + 1, min(i + 5, len(all_lines))):
                try:
                    v = float(all_lines[j])
                    if 0 < v <= 4:
                        result["교육실적_기재점수"] = v
                        break
                except ValueError:
                    pass
            break
        # 소계 없으면 첫 합계 점수(1.5~4) 탐색
        if in_edu:
            try:
                v = float(line)
                if 1.5 <= v <= 4.0:
                    result["교육실적_기재점수"] = v
                    break
            except ValueError:
                pass

    return result


def extract_overlap(doc):
    """업무중첩도 추출 (양식2-10)
    - 상주감리원: '실격' 키워드 탐지 → 없으면 6점 만점, 있으면 0점
    - 비상주감리원: '감리용역' 출현 횟수(비상주 섹션) → NONRESIDENT_OVERLAP_SCORE 적용
    """
    result = {
        "상주_실격": False,
        "상주_기재점수": 0.0,
        "상주_계산점수": 0.0,
        "비상주_현장수": 0,
        "비상주_기재점수": 0.0,
        "비상주_계산점수": 0.0,
        "업무중첩_기재점수": 0.0,
        "page": -1,
    }

    overlap_start = max(int(doc.page_count * 0.55), 40)
    page_num = smart_find_page(doc, ['양식2-10', '업무중첩도', '배치현황'],
                                range(overlap_start, doc.page_count), "업무중첩도")
    if page_num < 0:
        return result

    result["page"] = page_num
    text = doc[page_num].get_text()
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # 실격 탐지 (상주감리원 관련 컨텍스트)
    result["상주_실격"] = '실격' in text
    result["상주_계산점수"] = 0.0 if result["상주_실격"] else 6.0

    # 상주/비상주 기재점수 추출
    # 페이지 상단 패턴: 감리원이름들 → '6','6' → 배치기간들 → '4','4' → '10','10'
    for i, line in enumerate(lines):
        if line == '6' and i + 1 < len(lines) and lines[i + 1] == '6':
            result["상주_기재점수"] = 6.0
            break
    for i, line in enumerate(lines):
        if line == '4' and i + 1 < len(lines) and lines[i + 1] == '4':
            result["비상주_기재점수"] = 4.0
            break

    # 비상주감리원 현장 수: 하단 현장배치현황 섹션에서 '용역' 포함 라인 카운트
    # PDF 컬럼 순서로 추출되어 현장명이 '비상주감리원' 라벨보다 먼저 나올 수 있음
    in_detail = False
    site_count = 0
    for line in lines:
        if '배치현황' in line or '현장명' in line:
            in_detail = True
        if in_detail and '용역' in line and len(line) > 3:
            site_count += 1

    result["비상주_현장수"] = site_count
    result["비상주_계산점수"] = float(
        NONRESIDENT_OVERLAP_SCORE.get(min(site_count, 9), 0))
    result["업무중첩_기재점수"] = result["상주_기재점수"] + result["비상주_기재점수"]

    print(f"  [업무중첩도] p{page_num+1}: 상주={'실격' if result['상주_실격'] else '정상'}({result['상주_계산점수']}점), "
          f"비상주={site_count}개소→{result['비상주_계산점수']}점")
    return result


def extract_replacement(doc):
    """교체빈도 추출 (양식2-11)
    - 가. 감리업체: '%' 라인 직전 숫자 = 교체빈도율 → REPLACEMENT_RATE_SCORE
    - 나. 참여감리원: 교체일(날짜) 카운트 → 상주×0.5 + 비상주×0.1 감점
    """
    result = {
        "업체_교체율": 0.0,
        "업체_기재점수": 0.0,
        "업체_계산점수": 0.0,
        "감리원_상주교체수": 0,
        "감리원_비상주교체수": 0,
        "감리원_기재점수": 0.0,
        "감리원_계산점수": 0.0,
        "교체빈도_기재점수": 0.0,
        "page": -1,
    }

    repl_start = max(int(doc.page_count * 0.6), 50)
    page_num = smart_find_page(doc, ['양식2-11', '교체빈도율', '배치감리원수'],
                                range(repl_start, doc.page_count), "교체빈도")
    if page_num < 0:
        return result

    result["page"] = page_num
    text = doc[page_num].get_text()
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # 가. 감리업체 교체빈도율: '%' 직전 float 값
    for i, line in enumerate(lines):
        if line == '%' and i > 0:
            try:
                result["업체_교체율"] = float(lines[i - 1])
                break
            except ValueError:
                pass
    result["업체_계산점수"] = lookup_score(result["업체_교체율"], REPLACEMENT_RATE_SCORE)

    # 업체 기재점수: '가. 감리업체' 섹션 이후 첫 점수값 (0 < v <= 2.5)
    in_업체 = False
    for line in lines:
        if '가.' in line and '감리업체' in line:
            in_업체 = True
            continue
        if in_업체:
            try:
                v = float(line)
                if 0 < v <= 2.5:
                    result["업체_기재점수"] = v
                    break
            except ValueError:
                pass

    # 나. 참여감리원 교체건수: 날짜 패턴(교체일) 카운트
    date_pat = re.compile(r'\d{4}[-./]\d{2}[-./]\d{2}')
    in_participant = False
    in_nonresident = False
    상주_교체 = 0
    비상주_교체 = 0

    for line in lines:
        if '나.' in line and '참여감리원' in line:
            in_participant = True
        if in_participant and '비상주' in line:
            in_nonresident = True
        if in_participant and date_pat.match(line):
            if in_nonresident:
                비상주_교체 += 1
            else:
                상주_교체 += 1

    result["감리원_상주교체수"] = 상주_교체
    result["감리원_비상주교체수"] = 비상주_교체
    result["감리원_계산점수"] = round(
        max(0.0, 2.5 - 상주_교체 * 0.5 - 비상주_교체 * 0.1), 2)

    # 감리원 기재점수: '합계' 이후 첫 점수값
    for i, line in enumerate(lines):
        if '합' in line and '계' in line and i + 1 < len(lines):
            try:
                result["감리원_기재점수"] = float(lines[i + 1])
                break
            except ValueError:
                pass

    result["교체빈도_기재점수"] = result["업체_기재점수"] + result["감리원_기재점수"]
    print(f"  [교체빈도] p{page_num+1}: 업체={result['업체_교체율']:.2f}%→{result['업체_계산점수']}점, "
          f"감리원=상주{상주_교체}/비상주{비상주_교체}건→{result['감리원_계산점수']}점")
    return result


def extract_keea_certificates(doc):
    """keea.or.kr 발급 확인서에서 발급번호 추출 (OCR 사용)

    스캔 페이지에서 상단 10-22% 영역을 2x 줌 OCR하여
    'XXX-XX-XXXXXXXX-XXXX-XXXXX' 형식의 발급번호를 추출한다.
    감리원 수(보통 3명)만큼 찾으면 조기 종료한다.
    """
    certs = []
    seen = set()
    pat = re.compile(r'\d{2,4}-\d{2}-\d{7,10}-\d{3,5}-\d{3,6}')
    name_pat = re.compile(r'^[가-힣]{2,4}$')

    # 양식2-11(교체빈도) 직후 ~ +8페이지 범위에서 탐색
    search_start = max(int(doc.page_count * 0.6), 50)
    start = smart_find_page(doc, ['양식2-11', '교체빈도율'], range(search_start, doc.page_count))
    if start < 0:
        start = doc.page_count - 15
    search_range = range(start + 1, min(start + 9, doc.page_count))

    reader = get_reader()
    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom (속도↑, 정확도 충분)

    for page_num in search_range:
        # 텍스트가 있는 페이지(양식)는 건너뜀
        native_text = doc[page_num].get_text().strip()
        if len(native_text) > 20:
            continue

        page = doc[page_num]
        rect = page.rect
        # 발급번호 영역: 페이지 상단 10~22%
        top_rect = fitz.Rect(rect.x0, rect.y0 + rect.height * 0.10,
                             rect.x1, rect.y0 + rect.height * 0.22)
        pix = page.get_pixmap(matrix=mat, clip=top_rect)
        img_data = pix.tobytes('png')
        del pix

        try:
            results = reader.readtext(img_data, batch_size=1)
        except Exception:
            del img_data
            continue
        del img_data

        texts = [t for _, t, _ in results]
        all_text = ' '.join(texts)
        matches = pat.findall(all_text)
        if not matches:
            continue

        issue_no = matches[0]
        if issue_no in seen:
            continue
        seen.add(issue_no)

        # 성명 추출: 발급번호 라인 이후 첫 한글 이름 (2-4글자)
        name = ""
        found_no = False
        for t in texts:
            if pat.search(t):
                found_no = True
                continue
            if found_no and name_pat.match(t):
                name = t
                break

        certs.append({
            "발급번호": issue_no,
            "성명": name,
            "page": page_num,
        })
        print(f"    [KEEA] p{page_num+1}: {name} 발급번호={issue_no}")

        # 감리원 3명분 찾으면 조기 종료
        if len(certs) >= 3:
            break

    gc.collect()
    return certs



def extract_sanctions(doc):
    """제재내역 (영업정지) 추출 - PyMuPDF 텍스트 사용"""
    result = {
        "업체_영업정지": "없음",
        "감리원_자격정지": "없음",
        "감점": 0,
        "page": -1,
    }

    # 제재내역은 보통 재정상태와 같은 페이지 또는 근처
    page_num = smart_find_page(doc, ['제재내역', '영업정지', '이하여백'],
                                range(25, min(int(doc.page_count * 0.5), doc.page_count)), "제재내역")
    if page_num < 0:
        page_num = smart_find_page(doc, ['제재', '영업정지', '벌점'],
                                    range(20, min(int(doc.page_count * 0.6), doc.page_count)), "제재내역(확장)")

    if page_num >= 0:
        result["page"] = page_num + 1
        text = get_page_text(doc, page_num)

        if '이하여백' in text or '해당없음' in text or '해당 없음' in text:
            pass  # 제재내역 없음 확인
        else:
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            for line in lines:
                # 영업정지 기간 추출
                m = re.search(r'영업정지\s*[:：]?\s*(\d+)\s*[개월일]', line)
                if m:
                    result["업체_영업정지"] = f"{m.group(1)}개월"
                    months = int(m.group(1))
                    result["감점"] += months * 0.2  # 1월당 -0.2점

                # 자격정지 추출
                m = re.search(r'자격정지\s*[:：]?\s*(\d+)\s*[개월일]', line)
                if m:
                    result["감리원_자격정지"] = f"{m.group(1)}개월"
                    months = int(m.group(1))
                    result["감점"] += (months // 3) * 0.2  # 3월당 -0.2점

                # 부실벌점 추출
                m = re.search(r'벌점\s*[:：]?\s*(\d+\.?\d*)', line)
                if m:
                    result["감점"] += float(m.group(1))

    return result


def check_disqualification(personnel, career):
    """실격 여부 확인"""
    reasons = []

    # 1. 등급 미달 확인 (입찰공고에서 요구하는 등급)
    chief_grade = personnel["책임감리원"]["등급"]
    if chief_grade and chief_grade not in ["특급", "고급", "중급"]:
        reasons.append(f"책임감리원 등급 미달({chief_grade})")

    # 비상주감리원 고급 미만 시 실격
    nonres_grade = personnel["비상주감리원"]["등급"]
    if nonres_grade and nonres_grade not in ["특급", "고급"]:
        reasons.append(f"비상주감리원 등급 미달({nonres_grade})")

    return reasons


def get_cert_bonus(cert_name):
    """자격증 가점 계산"""
    if not cert_name:
        return 0

    if '기술사' in cert_name:
        return 1.0
    elif '기능장' in cert_name:
        return 0.8
    elif '기사' in cert_name and '산업' not in cert_name:
        return 0.7
    elif '산업기사' in cert_name:
        return 0.5
    return 0


def lookup_score(value, score_table):
    """임계값 기반 점수표 조회 (내림차순 임계값 테이블)
    score_table: [(threshold, score), ...] — 높은 임계값 순서로 정렬
    value >= threshold 이면 해당 점수 반환
    """
    for threshold, score in score_table:
        if value >= threshold:
            return score
    return score_table[-1][1]


def lookup_career_score(grade, years, table):
    """등급+경력년수로 경력점수 조회
    table 형식: {"특급": [(임계_미만, 점수), ...], ...}
      - (1, 4.0): years < 1 이면 4.0점
      - (0, 10.0): 0은 최고구간 표시 → 10.0점
    """
    if grade not in table:
        return 0.0
    for threshold, score in table[grade]:
        if threshold == 0:
            return score  # 최고 구간 (이상 없음)
        if years < threshold:
            return score
    return table[grade][-1][1]


def _dev_score_from_ratio(item_type, ratio_pct):
    """기재된 기간비율%와 타입으로 TECH_DEV_SCORE 기준 점수 반환.
    ratio_pct: 60 / 80 / 100 (기간비율, 정수)
    """
    table = TECH_DEV_SCORE.get(item_type, {})
    for key in sorted(table.keys(), reverse=True):
        if ratio_pct >= key - 1:   # 반올림 오차 1 허용
            return table[key]
    return 0.0


def calculate_scores_by_criteria(pdf_data):
    """심사기준표를 적용하여 점수를 독립 계산한다.

    pdf_data: analyze_company()가 반환한 중간 결과 dict
    반환: {
        "책임_전기경력_계산점수": float,
        "책임_전기경력_근거": str,
        "책임_배전경력1_계산점수": float,
        ...
    }
    """
    calc = {}
    cost_tier = pdf_data.get("cost_tier", COST_TIER)

    # ── 1. 책임감리원 ──
    grade = pdf_data.get("책임_등급", "")

    # 1-1. 전기분야 경력 (등급 무관, 공사비만)
    elec_years = pdf_data.get("책임_전기경력_년", 0) or 0
    s, b = calc_career_score(elec_years, "책임_전기", cost_tier=cost_tier)
    calc["책임_전기경력_계산점수"] = s
    calc["책임_전기경력_근거"] = b

    # 1-2. 참여분야 경력1 (설계·시공·감리 등 전체)
    field1_years = pdf_data.get("책임_배전경력1_년", 0) or 0
    s, b = calc_career_score(field1_years, "책임_참여1", grade, cost_tier)
    calc["책임_배전경력1_계산점수"] = s
    calc["책임_배전경력1_근거"] = b

    # 1-3. 참여분야 경력2 (감독·감리만)
    field2_years = pdf_data.get("책임_배전경력2_년", 0) or 0
    s, b = calc_career_score(field2_years, "책임_참여2", grade, cost_tier)
    calc["책임_배전경력2_계산점수"] = s
    calc["책임_배전경력2_근거"] = b

    # ── 2. 보조감리원 ──
    asst_grade = pdf_data.get("보조_등급", "")
    asst_elec_years = pdf_data.get("보조_전기경력_년", 0) or 0
    if asst_elec_years > 0 or pdf_data.get("보조_전기경력_점수", 0) > 0:
        s, b = calc_career_score(asst_elec_years, "보조_전기", asst_grade, cost_tier)
        calc["보조_전기경력_계산점수"] = s
        calc["보조_전기경력_근거"] = b

        asst_f1_years = pdf_data.get("보조_참여분야1_년", 0) or asst_elec_years
        s, b = calc_career_score(asst_f1_years, "보조_참여1", asst_grade, cost_tier)
        calc["보조_참여1_계산점수"] = s
        calc["보조_참여1_근거"] = b

        asst_f2_years = pdf_data.get("보조_참여분야2_년", 0) or asst_elec_years
        s, b = calc_career_score(asst_f2_years, "보조_참여2", asst_grade, cost_tier)
        calc["보조_참여2_계산점수"] = s
        calc["보조_참여2_근거"] = b

    # ── 3. 비상주감리원 ──
    nonres_grade = pdf_data.get("비상주_등급", "")
    # 등급 점수
    if nonres_grade in NONRESIDENT_GRADE_SCORE:
        calc["비상주_등급_계산점수"] = NONRESIDENT_GRADE_SCORE[nonres_grade]
        calc["비상주_등급_근거"] = nonres_grade
    elif nonres_grade:
        calc["비상주_등급_계산점수"] = 0.0
        calc["비상주_등급_근거"] = f"{nonres_grade} (고급 미만→실격)"

    # 전기분야 경력
    nonres_elec_years = pdf_data.get("비상주_전기경력_년", 0) or 0
    if nonres_elec_years > 0 or pdf_data.get("비상주_전기경력_점수", 0) > 0:
        s, b = calc_career_score(nonres_elec_years, "비상주_전기", nonres_grade, cost_tier)
        calc["비상주_전기경력_계산점수"] = s
        calc["비상주_전기경력_근거"] = b

    # 5. 유사용역 점수 (사정금액 대비 비율 기반)
    ratio = pdf_data.get("유사용역_비율", 0) or 0
    if ratio > 0:
        calc["유사용역_계산점수"] = lookup_score(ratio, SIMILAR_PROJECT_SCORE)
        calc["유사용역_근거"] = f"비율 {ratio}%"

    # 6. 책임감리원 자격증 가점
    cert = pdf_data.get("책임_자격증", "") or ""
    calc["가점_계산점수"] = get_cert_bonus(cert)
    calc["가점_근거"] = cert if cert else "해당없음"

    # 7. 기술개발 투자실적 (A÷B%)
    invest_ratio = pdf_data.get("기술투자_비율", 0) or 0
    if invest_ratio > 0:
        calc["기술투자_계산점수"] = lookup_score(invest_ratio, TECH_INVEST_SCORE)
        calc["기술투자_근거"] = f"A÷B={invest_ratio:.2f}%"

    # 8. 개발실적 (타입별 기간비율 기반 독립 합산, max 4점)
    dev_calc = pdf_data.get("개발실적_계산점수", None)
    dev_cnt = pdf_data.get("개발실적_항목수", 0)
    if dev_calc is not None and dev_cnt > 0:
        calc["개발실적_계산점수"] = dev_calc
        calc["개발실적_근거"] = f"{dev_cnt}건 합산(max4점)"

    # 8-1. 교육실적 (최근3년 전기공사감리 교육훈련 이수 시 2점)
    edu_score = pdf_data.get("교육실적_기재점수", 0)
    if edu_score > 0:
        calc["교육실적_계산점수"] = 2.0
        calc["교육실적_근거"] = "교육이수→2점"
    else:
        calc["교육실적_계산점수"] = 0.0
        calc["교육실적_근거"] = "교육미이수→0점"

    # 9. 상주감리원 중첩도
    overlap_raw = pdf_data.get("overlap_raw", {})
    if overlap_raw:
        상주_실격 = overlap_raw.get("상주_실격", False)
        calc["상주중첩_계산점수"] = 0.0 if 상주_실격 else 6.0
        calc["상주중첩_근거"] = "실격" if 상주_실격 else "실격없음→만점"

    # 10. 비상주감리원 중첩도
    site_count = overlap_raw.get("비상주_현장수", -1) if overlap_raw else -1
    if site_count >= 0:
        score = float(NONRESIDENT_OVERLAP_SCORE.get(min(site_count, 9), 0))
        calc["비상주중첩_계산점수"] = score
        calc["비상주중첩_근거"] = f"{site_count}개소→{score}점"

    # 11. 감리업체 교체빈도
    replacement_raw = pdf_data.get("replacement_raw", {})
    if replacement_raw:
        rate = replacement_raw.get("업체_교체율", 0)
        calc["업체교체_계산점수"] = lookup_score(rate, REPLACEMENT_RATE_SCORE)
        calc["업체교체_근거"] = f"교체율{rate:.2f}%→{calc['업체교체_계산점수']}점"

        # 12. 참여감리원 교체빈도
        상주 = replacement_raw.get("감리원_상주교체수", 0)
        비상주 = replacement_raw.get("감리원_비상주교체수", 0)
        score = round(max(0.0, 2.5 - 상주 * 0.5 - 비상주 * 0.1), 2)
        calc["감리원교체_계산점수"] = score
        calc["감리원교체_근거"] = f"상주{상주}건/비상주{비상주}건→{score}점"

    # 13. 부실벌점 (누계평균부실벌점 기준, KEEA 해당없음→감점0)
    penalty_avg = pdf_data.get("부실벌점_누계평균", None)
    keea_certs = pdf_data.get("keea_certs", [])
    if penalty_avg is not None and penalty_avg > 0:
        deduction = lookup_score(penalty_avg, PENALTY_DEDUCTION_SCORE)
        calc["부실벌점_계산점수"] = deduction
        calc["부실벌점_근거"] = f"누계평균{penalty_avg}점→{deduction}점"
    elif keea_certs:
        calc["부실벌점_계산점수"] = 0.0
        calc["부실벌점_근거"] = "KEEA확인서 해당없음→감점0"

    return calc


###############################################################################
# 메인 분석 함수
###############################################################################

def analyze_company(pdf_path, bidding_date=BIDDING_DATE, cost_tier=COST_TIER):
    """업체 PQ 제출 PDF 전체 분석"""
    print(f"\n{'='*60}")
    print(f"  배전감리 PQ 심사 데이터 추출")
    print(f"  대상 파일: {os.path.basename(pdf_path)}")
    print(f"  입찰공고일: {bidding_date}")
    print(f"{'='*60}\n")

    set_current_pdf(pdf_path)  # 파일 캐시 경로 설정
    doc = fitz.open(pdf_path)
    print(f"[INFO] PDF 열림: 총 {doc.page_count} 페이지")
    cache_dir = _get_cache_dir()
    if cache_dir and os.path.exists(cache_dir):
        cached_count = len([f for f in os.listdir(cache_dir) if f.endswith('.json')])
        if cached_count > 0:
            print(f"[INFO] 캐시 발견: {cached_count}페이지 OCR 결과 재사용")
    print()

    # 1. 업체명 추출
    print("[1/5] 업체명 추출 중...")
    company_name = extract_company_name(doc)
    print(f"  → 업체명: {company_name}")

    # 2. 종합득점표 추출
    print("[2/5] 종합득점표 추출 중...")
    summary = extract_summary_table(doc)
    print(f"  → 종합득점표 위치: p{summary['page']}")
    print(f"  → 참여감리원: {summary['참여감리원']}점, 유사용역: {summary['유사용역']}점")
    print(f"  → 기술개발: {summary['기술개발']}점, 업무중첩도: {summary['업무중첩도']}점")
    print(f"  → 교체빈도: {summary['교체빈도']}점, 가감점: {summary['가감점']}점")
    print(f"  → 업체제출 총점: {summary['총점']}")

    # 2.5 페이지 분류 (find_tables 기반 추출용)
    page_map = classify_pages(doc)
    print(f"  → 페이지 분류: {', '.join(f'{k}:p{v[0]+1}' for k,v in page_map.items() if v)}")

    # 3. 참여감리원 자격사항 추출
    print("[3/5] 참여감리원 자격사항 추출 중...")
    personnel = extract_personnel_v2(doc, page_map)
    if personnel is None:
        print("  → [v2 실패] 기존 방식으로 폴백")
        personnel = extract_personnel(doc)
    print(f"  → 책임감리원: {personnel['책임감리원']['성명']} / {personnel['책임감리원']['등급']}")
    if personnel["보조감리원"]:
        print(f"  → 보조감리원: {personnel['보조감리원'][0]['성명']}")
    print(f"  → 비상주감리원: {personnel['비상주감리원']['성명']} / {personnel['비상주감리원']['등급']}")

    # 4. 책임감리원 경력 추출
    print("[4/5] 책임감리원 경력 추출 중...")
    career = extract_career_v2(doc, page_map, "책임")
    if career is None:
        print("  → [v2 실패] 기존 방식으로 폴백")
        career = extract_career_summary(doc)
    print(f"  → 전기분야: {career['책임_전기분야_년']}년 → {career['책임_전기분야_점수']}점")
    print(f"  → 참여분야(전체): {career['책임_참여분야1_년']}년 → {career['책임_참여분야1_점수']}점")
    print(f"  → 참여분야(감독감리): {career['책임_참여분야2_년']}년 → {career['책임_참여분야2_점수']}점")

    # 4.5 보조감리원/비상주감리원 경력 추출
    print("[4.5/5] 보조/비상주감리원 경력 추출 중...")
    asst_career = extract_career_v2(doc, page_map, "보조")
    if asst_career is None:
        chief_page = career["page"] - 1 if career["page"] > 0 else 17
        asst_career = extract_asst_career(doc, chief_page)
    nonres_career = extract_career_v2(doc, page_map, "비상주")
    if nonres_career is None:
        chief_page = career["page"] - 1 if career["page"] > 0 else 17
        nonres_career = extract_nonres_career(doc, chief_page)

    # 5. 유사용역 실적 추출
    print("[5/5] 유사용역 수행실적 추출 중...")
    similar = extract_similar_project(doc)
    print(f"  → 적용금액: {similar['적용금액']:,}원")
    if similar["점수"] > 0:
        print(f"  → 평점: {similar['점수']}점")

    # 5.5 기술개발 세부증빙 추출 (양식2-9)
    print("[5.5/5] 기술개발 세부증빙 추출 중...")
    tech_dev = extract_tech_development(doc, bidding_date)
    if tech_dev["page"] > 0:
        print(f"  → p{tech_dev['page']}: 개발실적 {tech_dev['개발실적_항목수']}건 "
              f"→ {tech_dev['개발실적_계산점수']}점")
        if tech_dev["매출액_합계"] > 0:
            print(f"  → 투자비율 {tech_dev['투자비율']:.2f}% "
                  f"→ {tech_dev['기술투자_계산점수']}점")
    else:
        print("  → 기술개발 페이지 미발견")

    # 5.7 업무중첩도 세부증빙 추출 (양식2-10)
    print("[5.7/5] 업무중첩도 세부증빙 추출 중...")
    overlap = extract_overlap(doc)

    # 5.9 교체빈도 세부증빙 추출 (양식2-11)
    print("[5.9/5] 교체빈도 세부증빙 추출 중...")
    replacement = extract_replacement(doc)

    # 5.95 KEEA 확인서 발급번호 추출 (OCR)
    print("[5.95/5] KEEA 확인서 발급번호 OCR 추출 중...")
    keea_certs = extract_keea_certificates(doc)


    # 제재내역 추출 (간략)
    sanctions = extract_sanctions(doc)

    # 실격 여부 판정
    disq_reasons = check_disqualification(personnel, career)

    # 가점 계산
    cert_bonus = get_cert_bonus(personnel["책임감리원"]["자격증"])

    doc.close()

    # 결과 종합
    result = {
        "업체명": company_name,
        "실격여부": "실격" if disq_reasons else "적격",
        "실격사유": "; ".join(disq_reasons) if disq_reasons else "없음",
        "책임_성명": personnel["책임감리원"]["성명"],
        "책임_등급": personnel["책임감리원"]["등급"],
        "책임_자격증": personnel["책임감리원"]["자격증"],
        "책임_전기경력_개월": career["책임_전기분야_개월"],
        "책임_전기경력_년": career["책임_전기분야_년"],
        "책임_전기경력_점수": career["책임_전기분야_점수"],
        "책임_배전경력1_년": career["책임_참여분야1_년"],
        "책임_배전경력1_점수": career["책임_참여분야1_점수"],
        "책임_배전경력2_년": career["책임_참여분야2_년"],
        "책임_배전경력2_점수": career["책임_참여분야2_점수"],
        "보조1_성명": personnel["보조감리원"][0]["성명"] if personnel["보조감리원"] else "",
        "보조_등급": asst_career["보조_등급"] or (personnel["보조감리원"][0]["등급"] if personnel["보조감리원"] else ""),
        "보조_전기경력_년": asst_career["보조_전기분야_년"],
        "보조_전기경력_점수": asst_career["보조_전기분야_점수"],
        "보조_참여분야1_점수": asst_career["보조_참여분야1_점수"],
        "보조_참여분야2_점수": asst_career["보조_참여분야2_점수"],
        "비상주_성명": personnel["비상주감리원"]["성명"],
        "비상주_등급": personnel["비상주감리원"]["등급"],
        "비상주_등급_점수": nonres_career["비상주_등급_점수"],
        "비상주_전기경력_년": nonres_career["비상주_전기분야_년"],
        "비상주_전기경력_점수": nonres_career["비상주_전기분야_점수"],
        "유사용역_적용금액": similar["적용금액"],
        "유사용역_비율": similar.get("비율", 0),
        "유사용역_점수": summary["유사용역"],
        "기술개발_점수": summary["기술개발"],
        "업무중첩_점수": summary["업무중첩도"],
        "교체빈도_점수": summary["교체빈도"],
        "참여감리원_소계": summary["참여감리원"],
        "영업정지": sanctions["업체_영업정지"],
        "가점_자격증": cert_bonus,
        "업체제출_총점": summary["총점"],
        # 기술개발 세부증빙 (양식2-9 독립 추출)
        "기술투자_비율": tech_dev["투자비율"],
        "기술투자_투자액": tech_dev["투자액_합계"],
        "기술투자_매출액": tech_dev["매출액_합계"],
        "기술투자_계산점수": tech_dev["기술투자_계산점수"],
        "개발실적_계산점수": tech_dev["개발실적_계산점수"],
        "개발실적_항목수": tech_dev["개발실적_항목수"],
        "교육실적_기재점수": tech_dev["교육실적_기재점수"],
        "tech_dev_raw": tech_dev,
        # 업무중첩도 세부증빙 (양식2-10 독립 추출)
        "상주중첩_기재점수": overlap["상주_기재점수"],
        "비상주중첩_기재점수": overlap["비상주_기재점수"],
        "업무중첩_기재점수": overlap["업무중첩_기재점수"],
        "상주_실격": overlap["상주_실격"],
        "비상주_현장수": overlap["비상주_현장수"],
        "overlap_raw": overlap,
        # 교체빈도 세부증빙 (양식2-11 독립 추출)
        "업체교체_기재점수": replacement["업체_기재점수"],
        "감리원교체_기재점수": replacement["감리원_기재점수"],
        "교체빈도_기재점수": replacement["교체빈도_기재점수"],
        "replacement_raw": replacement,
        # KEEA 확인서 발급번호
        "keea_certs": keea_certs,
        # 공사비 구간
        "cost_tier": cost_tier,
        # 근거 페이지
        "근거페이지": {
            "종합득점표": summary["page"],
            "참여감리원": personnel["page"],
            "경력실적": career["page"],
            "유사용역": similar["page"],
            "제재내역": sanctions["page"],
            "기술개발": tech_dev["page"],
            "업무중첩도": overlap["page"],
            "교체빈도": replacement["page"],
        }
    }

    # 심사기준표 기반 점수 독립 계산
    result["criteria_scores"] = calculate_scores_by_criteria(result)

    return result


###############################################################################
# Excel 출력
###############################################################################

def export_to_excel(result, output_path):
    """분석 결과를 Excel로 출력"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PQ심사결과"

    # 스타일 설정
    header_font = Font(name='맑은 고딕', bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name='맑은 고딕', bold=True, size=11, color="FFFFFF")
    data_font = Font(name='맑은 고딕', size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws.merge_cells('A1:H1')
    ws['A1'] = '배전감리 PQ 심사 데이터 추출 결과'
    ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'입찰공고일: {BIDDING_DATE} | 추출일시: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='맑은 고딕', size=9, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')

    # 헤더 (3행)
    headers = [
        '업체명', '실격여부\n(사유)', '책임감리원\n성명/등급',
        '책임_전기경력\n(년/점수)', '책임_배전경력\n(년/점수)',
        '보조/비상주\n성명', '유사용역\n최종실적액',
        '근거페이지\n(PDF)'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 데이터 (4행)
    row = 5

    # 1. 업체명
    ws.cell(row=row, column=1, value=result["업체명"]).font = data_font

    # 2. 실격여부
    disq_text = result["실격여부"]
    if result["실격사유"] != "없음":
        disq_text += f"\n({result['실격사유']})"
    cell = ws.cell(row=row, column=2, value=disq_text)
    cell.font = data_font
    if result["실격여부"] == "실격":
        cell.font = Font(name='맑은 고딕', size=10, color="FF0000", bold=True)

    # 3. 책임감리원 성명/등급
    ws.cell(row=row, column=3,
            value=f"{result['책임_성명']}\n{result['책임_등급']}").font = data_font

    # 4. 전기경력
    ws.cell(row=row, column=4,
            value=f"{result['책임_전기경력_년']}년\n→ {result['책임_전기경력_점수']}점").font = data_font

    # 5. 배전경력
    ws.cell(row=row, column=5,
            value=f"전체: {result['책임_배전경력1_년']}년 → {result['책임_배전경력1_점수']}점\n"
                  f"감독감리: {result['책임_배전경력2_년']}년 → {result['책임_배전경력2_점수']}점").font = data_font

    # 6. 보조/비상주
    ws.cell(row=row, column=6,
            value=f"보조: {result['보조1_성명']}\n비상주: {result['비상주_성명']}({result['비상주_등급']})").font = data_font

    # 7. 유사용역 실적액
    ws.cell(row=row, column=7,
            value=f"{result['유사용역_적용금액']:,}원").font = data_font

    # 8. 근거페이지
    pages = result["근거페이지"]
    ws.cell(row=row, column=8,
            value=f"종합득점표: p{pages['종합득점표']}\n"
                  f"참여감리원: p{pages['참여감리원']}\n"
                  f"경력실적: p{pages['경력실적']}").font = data_font

    # 스타일 적용
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

    # 열 너비 조정
    widths = [18, 15, 15, 18, 22, 18, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # 행 높이
    ws.row_height = 60
    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 100

    # 상세 데이터 시트
    ws2 = wb.create_sheet("상세데이터")
    ws2['A1'] = '항목'
    ws2['B1'] = '값'
    ws2['C1'] = '비고'
    ws2['A1'].font = header_font
    ws2['B1'].font = header_font
    ws2['C1'].font = header_font

    detail_rows = [
        ('업체명', result['업체명'], ''),
        ('실격여부', result['실격여부'], result['실격사유']),
        ('---참여감리원---', '', ''),
        ('책임감리원 성명', result['책임_성명'], ''),
        ('책임감리원 등급', result['책임_등급'], ''),
        ('책임감리원 자격증', result['책임_자격증'], f"가점: {result['가점_자격증']}점"),
        ('책임_전기경력(개월)', result['책임_전기경력_개월'], ''),
        ('책임_전기경력(년)', result['책임_전기경력_년'], f"절사후 → {result['책임_전기경력_점수']}점"),
        ('책임_배전경력1(년)', result['책임_배전경력1_년'], f"전체 → {result['책임_배전경력1_점수']}점"),
        ('책임_배전경력2(년)', result['책임_배전경력2_년'], f"감독감리 → {result['책임_배전경력2_점수']}점"),
        ('보조감리원 성명', result['보조1_성명'], ''),
        ('비상주감리원 성명', result['비상주_성명'], ''),
        ('비상주감리원 등급', result['비상주_등급'], ''),
        ('---유사용역---', '', ''),
        ('유사용역 적용금액', f"{result['유사용역_적용금액']:,}원", ''),
        ('영업정지', result['영업정지'], ''),
        ('---합계---', '', ''),
        ('업체제출 총점', result['업체제출_총점'], '100점 만점 기준'),
    ]

    for i, (item, val, note) in enumerate(detail_rows, 2):
        ws2.cell(row=i, column=1, value=item).font = data_font
        ws2.cell(row=i, column=2, value=str(val)).font = data_font
        ws2.cell(row=i, column=3, value=note).font = data_font

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 30

    wb.save(output_path)
    print(f"\n[OUTPUT] Excel 저장 완료: {output_path}")


###############################################################################
# CSV 출력
###############################################################################

def print_csv_result(result):
    """CSV 형태로 콘솔 출력"""
    print(f"\n{'='*60}")
    print("  CSV 출력 결과")
    print(f"{'='*60}")
    print(f"1. [업체명]: {result['업체명']}")
    print(f"2. [실격여부]: {result['실격여부']} ({result['실격사유']})")
    print(f"3. [책임_성명/등급]: {result['책임_성명']} / {result['책임_등급']}")
    print(f"4. [책임_전기경력_일수]: {result['책임_전기경력_개월']}개월 ({result['책임_전기경력_년']}년) → {result['책임_전기경력_점수']}점")
    print(f"   [책임_배전경력_일수]: 전체 {result['책임_배전경력1_년']}년 → {result['책임_배전경력1_점수']}점 / 감독감리 {result['책임_배전경력2_년']}년 → {result['책임_배전경력2_점수']}점")
    print(f"5. [보조1_성명]: {result['보조1_성명']}")
    print(f"   [비상주_성명/등급]: {result['비상주_성명']} / {result['비상주_등급']}")
    print(f"6. [유사용역_최종실적액]: {result['유사용역_적용금액']:,}원")
    print(f"7. [영업정지]: {result['영업정지']}")
    print(f"8. [가점_자격증]: {result['가점_자격증']}점 ({result['책임_자격증']})")

    pages = result["근거페이지"]
    print(f"9. [근거페이지]: 종합득점표 p{pages['종합득점표']}, 참여감리원 p{pages['참여감리원']}, "
          f"경력실적 p{pages['경력실적']}")

    print(f"\n[업체제출 총점]: {result['업체제출_총점']}점")


###############################################################################
# 메인 실행
###############################################################################

if __name__ == "__main__":
    # 기본 경로
    pdf_path = r"C:\Users\Admin\Desktop\project\pq\증빙\(주)성문기술단(505-81-31517)-복사.pdf"

    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    if len(sys.argv) > 2:
        BIDDING_DATE = sys.argv[2]

    if not os.path.exists(pdf_path):
        print(f"[ERROR] 파일을 찾을 수 없습니다: {pdf_path}")
        sys.exit(1)

    # 분석 실행
    result = analyze_company(pdf_path, BIDDING_DATE)

    # CSV 형태 콘솔 출력
    print_csv_result(result)

    # Excel 출력
    output_dir = os.path.dirname(pdf_path) or "."
    output_path = os.path.join(os.path.dirname(os.path.dirname(pdf_path)),
                               "PQ심사결과_" + result["업체명"].replace("(", "").replace(")", "") + ".xlsx")
    export_to_excel(result, output_path)

    print(f"\n{'='*60}")
    print("  분석 완료!")
    print(f"{'='*60}")
